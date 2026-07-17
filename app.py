import os
import csv
import io
import hmac
import hashlib
import base64
import logging
import time
import threading
from collections import defaultdict
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from flask import Flask, Response, request, jsonify, redirect
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from requests_aws4auth import AWS4Auth
import openpyxl
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import pytz
import sys
sys.stdout.reconfigure(line_buffering=True)

load_dotenv()
requests.adapters.DEFAULT_RETRIES = 3

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

app = Flask(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# SHARED CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
AIRTABLE_URL    = "https://api.airtable.com/v0"
REQUEST_TIMEOUT = 30


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — DELIVERY TRACKER (professionalcourier.ae + Shopify)
# ══════════════════════════════════════════════════════════════════════════════

SHOPIFY_STORE        = os.getenv("SHOPIFY_STORE")
SHOPIFY_ACCESS_TOKEN = os.getenv("SHOPIFY_ACCESS_TOKEN")
API_VERSION          = os.getenv("SHOPIFY_API_VERSION", "2024-04")

# for API access token
CLIENT_ID     = os.getenv("SHOPIFY_CLIENT_ID")
CLIENT_SECRET = os.getenv("SHOPIFY_CLIENT_SECRET")
REDIRECT_URI  = os.getenv("SHOPIFY_REDIRECT_URI")  # e.g. https://your-render-url.com/auth/callback

HEADERS = {
    "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
    "Content-Type": "application/json",
}

def shopify(path):
    return f"https://{SHOPIFY_STORE}/admin/api/{API_VERSION}{path}"

# ── Startup banner ─────────────────────────────────────────────────────────────
print("=" * 60, flush=True)
print("  DELIVERY TRACKER — STARTING", flush=True)
print(f"  Store : {SHOPIFY_STORE}", flush=True)
print(f"  Token : {'SET ✓' if SHOPIFY_ACCESS_TOKEN else 'MISSING ✗'}", flush=True)
print("=" * 60, flush=True)


# ── 1. Fetch only orders that need delivery check ─────────────────────────────
def get_orders_needing_delivery_check():
    print("\n[SHOPIFY] Fetching orders where Delivery Status = Tracking added...", flush=True)
    all_orders = []
    url    = shopify("/orders.json")
    params = {
        "fulfillment_status": "shipped",
        "status":             "any",
        "limit":              250,
    }

    page = 0
    while url:
        page += 1
        print(f"  → Page {page}: GET {url}", flush=True)
        r = requests.get(url, headers=HEADERS, params=params, timeout=30)
        r.raise_for_status()

        orders = r.json().get("orders", [])
        print(f"    Fetched {len(orders)} orders from Shopify", flush=True)

        for order in orders:
            needs_check = any(
                (ful.get("shipment_status") or "") != "delivered"
                and (ful.get("tracking_company") or "").strip().lower() == "other"
                and (ful.get("tracking_number") or "").strip()
                for ful in order.get("fulfillments", [])
            )
            if needs_check:
                all_orders.append(order)

        print(f"    {len(all_orders)} orders need delivery check so far", flush=True)

        link   = r.headers.get("Link", "")
        url    = None
        params = None
        if 'rel="next"' in link:
            for part in link.split(","):
                if 'rel="next"' in part:
                    url = part.split(";")[0].strip().strip("<>")
                    break

    print(f"[SHOPIFY] Orders needing delivery check: {len(all_orders)}\n", flush=True)
    return all_orders


# ── 2. Mark Delivery Status = Delivered in Shopify ───────────────────────────

def mark_delivered(order_id, fulfillment_id):
    r = requests.post(
        shopify(f"/orders/{order_id}/fulfillments/{fulfillment_id}/events.json"),
        headers=HEADERS,
        json={"event": {"status": "delivered"}},
        timeout=15,
    )
    r.raise_for_status()
    return r.json().get("fulfillment_event", {})


# ── 3. Scrape professionalcourier.ae ─────────────────────────────────────────
# Confirmed from browser DevTools:
#   Form action : https://professionalcourier.ae/tracking  (POST)
#   Field name  : trackno
# Flow: GET first to obtain session cookie → POST with trackno=AWB

def check_courier(tracking_number: str) -> dict:
    TRACKING_URL = "https://professionalcourier.ae/tracking"
    print(f"    [COURIER] Checking AWB {tracking_number}...", flush=True)

    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })

    # ── Step 1: GET the page to obtain session cookies ────────────────────────
    try:
        get_resp = session.get(TRACKING_URL, timeout=20)
        get_resp.raise_for_status()
        print(f"    [COURIER] GET OK — {len(get_resp.text)} chars | "
              f"cookies: {list(session.cookies.keys())}", flush=True)
    except Exception as e:
        print(f"    [COURIER] ✗ GET failed: {e}", flush=True)
        return {"is_delivered": False, "status": "unreachable", "error": str(e)}

    # ── Step 2: POST with correct field name "trackno" ────────────────────────
    try:
        resp = session.post(
            TRACKING_URL,
            data={"trackno": tracking_number},
            headers={
                "Referer":      TRACKING_URL,
                "Origin":       "https://professionalcourier.ae",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            timeout=20,
        )
        resp.raise_for_status()
        print(f"    [COURIER] POST {resp.status_code} — {len(resp.text)} chars", flush=True)
    except Exception as e:
        print(f"    [COURIER] ✗ POST failed: {e}", flush=True)
        return {"is_delivered": False, "status": "post_failed", "error": str(e)}

    result_soup = BeautifulSoup(resp.text, "html.parser")
    page_text   = result_soup.get_text(" ", strip=True)

    # ── Step 3: Verify tracking number appears in result ─────────────────────
    if tracking_number not in page_text:
        print(f"    [COURIER] ✗ Tracking number not found in result", flush=True)
        # Print snippet for debugging
        print(f"    [COURIER] Page snippet: {page_text[:200]}", flush=True)
        return {"is_delivered": False, "status": "not_found"}

    print(f"    [COURIER] ✓ Tracking number found in result", flush=True)

    # ── Step 4: Find "Current Status" column in summary table ─────────────────
    # Table structure: From | To | Current Status | Current Activity
    status_text = ""
    for table in result_soup.find_all("table"):
        headers = [th.get_text(strip=True).lower() for th in table.find_all("th")]
        print(f"    [COURIER] Table headers: {headers}", flush=True)
        if "current status" in headers:
            try:
                si   = headers.index("current status")
                rows = table.find_all("tr")
                if len(rows) > 1:
                    cells = rows[1].find_all("td")
                    if cells and si < len(cells):
                        status_text = cells[si].get_text(strip=True)
                        print(f"    [COURIER] Current Status: '{status_text}'", flush=True)
            except (ValueError, IndexError) as e:
                print(f"    [COURIER] Table parse error: {e}", flush=True)
            break

    # ── Step 5: Fallback — scan text near tracking number only ────────────────
    if not status_text:
        idx = page_text.find(tracking_number)
        if idx != -1:
            nearby = page_text[idx: idx + 400].lower()
            for k in ["delivered", "out for delivery", "in transit",
                      "dispatched", "picked up", "processing", "pending"]:
                if k in nearby:
                    status_text = k.title()
                    print(f"    [COURIER] Fallback status: '{status_text}'", flush=True)
                    break

    # ── Step 6: Exact match only — never whole-page match ────────────────────
    is_delivered = status_text.strip().lower() in (
        "delivered", "delivery complete", "successfully delivered"
    )

    print(f"    [COURIER] Final → status='{status_text}' is_delivered={is_delivered}", flush=True)
    return {"is_delivered": is_delivered, "status": status_text or "unknown"}


# ── Main logic ────────────────────────────────────────────────────────────────

def run_tracking():
    print("\n" + "=" * 60, flush=True)
    print("  RUN TRACKING STARTED", flush=True)
    print("=" * 60, flush=True)

    summary = {
        "checked": 0, "updated": 0,
        "errors":  0, "skipped": 0,
        "details": []
    }

    try:
        orders = get_orders_needing_delivery_check()
    except Exception as e:
        print(f"[ERROR] Failed to fetch orders: {e}", flush=True)
        summary["errors"] += 1
        return summary

    print(f"[PROCESSING] {len(orders)} orders to check...\n", flush=True)

    for order in orders:
        order_number = order.get("order_number") or order.get("name")
        order_id     = order["id"]

        for ful in order.get("fulfillments", []):
            ful_id           = ful["id"]
            tracking_number  = (ful.get("tracking_number") or "").strip()
            tracking_company = (ful.get("tracking_company") or "").strip()
            shipment_status  = (ful.get("shipment_status") or "").lower()

            print(f"\n  Order #{order_number} | AWB: {tracking_number} | "
                  f"Carrier: {tracking_company} | Status: {shipment_status}", flush=True)

            detail = {
                "order":   order_number,
                "awb":     tracking_number,
                "carrier": tracking_company,
                "status":  shipment_status,
                "action":  None,
            }

            if shipment_status == "delivered":
                msg = "skip — already delivered"
                print(f"  → {msg}", flush=True)
                detail["action"] = msg
                summary["skipped"] += 1
                summary["details"].append(detail)
                continue

            if tracking_company.lower() != "other":
                msg = f"skip — carrier is '{tracking_company}' not 'Other'"
                print(f"  → {msg}", flush=True)
                detail["action"] = msg
                summary["skipped"] += 1
                summary["details"].append(detail)
                continue

            if not tracking_number:
                msg = "skip — no tracking number"
                print(f"  → {msg}", flush=True)
                detail["action"] = msg
                summary["skipped"] += 1
                summary["details"].append(detail)
                continue

            print(f"  → ✓ Conditions met — checking professionalcourier.ae...", flush=True)
            summary["checked"] += 1

            courier = check_courier(tracking_number)

            if courier.get("error"):
                msg = f"error: {courier['error']}"
                print(f"  → ✗ {msg}", flush=True)
                detail["action"] = msg
                summary["errors"] += 1
                summary["details"].append(detail)
                continue

            if courier["is_delivered"]:
                try:
                    mark_delivered(order_id, ful_id)
                    msg = "✅ MARKED DELIVERED in Shopify"
                    print(f"  → {msg}", flush=True)
                    detail["action"] = msg
                    summary["updated"] += 1
                except Exception as e:
                    msg = f"Shopify update failed: {e}"
                    print(f"  → ✗ {msg}", flush=True)
                    detail["action"] = msg
                    summary["errors"] += 1
            else:
                msg = f"not delivered yet (courier: {courier['status']})"
                print(f"  → {msg}", flush=True)
                detail["action"] = msg

            summary["details"].append(detail)

    print("\n" + "=" * 60, flush=True)
    print(f"  RUN COMPLETE", flush=True)
    print(f"  Checked : {summary['checked']}", flush=True)
    print(f"  Updated : {summary['updated']}", flush=True)
    print(f"  Skipped : {summary['skipped']}", flush=True)
    print(f"  Errors  : {summary['errors']}", flush=True)
    print("=" * 60 + "\n", flush=True)

    return summary


# ── Routes (Section 1) ────────────────────────────────────────────────────────

@app.route("/check-tracking", methods=["POST", "GET"])
def check_tracking():
    """
    Called by Shopify Flow at 9am and 6pm IST.
    Responds immediately to avoid Flow's 30-second timeout.
    Tracking runs in background — check Render logs for results.
    """
    import threading
    print(f"\n>>> /check-tracking triggered — starting background job", flush=True)
    thread = threading.Thread(target=run_tracking, daemon=True)
    thread.start()
    return jsonify({"ok": True, "message": "Tracking job started in background"}), 200


@app.route("/delivery/health", methods=["GET"])
# NOTE: renamed from /health to /delivery/health to avoid conflict with Section 4
def delivery_health():
    """Ping this every 14 min from UptimeRobot to keep Render awake."""
    print(">>> GET /delivery/health — OK", flush=True)
    return jsonify({"status": "ok", "store": SHOPIFY_STORE}), 200


@app.route("/", methods=["GET"])
def index():
    return jsonify({
        "service":   "Delivery Sync — Fragrant Souq",
        "endpoints": {
            "POST /check-tracking": "Run tracking (called by Shopify Flow)",
            "GET  /delivery/health": "Health check for uptime monitors",
        }
    }), 200


#  API access token
@app.route("/auth", methods=["GET"])
def auth():
    shop = request.args.get("shop", SHOPIFY_STORE)
    scopes = "read_orders,write_orders,read_all_orders,read_fulfillments,write_fulfillments,read_customers,write_customers"
    auth_url = (
        f"https://{shop}/admin/oauth/authorize"
        f"?client_id={CLIENT_ID}"
        f"&scope={scopes}"
        f"&redirect_uri={REDIRECT_URI}"
    )
    return redirect(auth_url)


@app.route("/auth/callback", methods=["GET"])
def auth_callback():
    code = request.args.get("code")
    shop = request.args.get("shop")  # ← use shop from callback params

    if not code:
        return "No code received", 400
    if not shop:
        return "No shop received", 400

    token_url = f"https://{shop}/admin/oauth/access_token"
    response = requests.post(token_url, json={
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "code":          code
    })

    print(f"🔑 Token exchange response: {response.status_code} — {response.text}", flush=True)

    token_data = response.json()
    access_token = token_data.get("access_token")
    print(f"🔑 NEW ACCESS TOKEN: {access_token}", flush=True)

    return jsonify({
        "access_token": access_token,
        "shop": shop,
        "message": "Copy this token and update SHOPIFY_TOKEN in your airtable service on Render!"
    })


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — AMAZON → AIRTABLE SYNC
# ══════════════════════════════════════════════════════════════════════════════

AIRTABLE_TOKEN              = os.getenv("AIRTABLE_TOKEN")
BASE_ID                     = os.getenv("BASE_ID") or os.getenv("AIRTABLE_BASE_ID")
CUSTOMERS_TABLE_ID          = os.getenv("CUSTOMERS_TABLE")
ORDER_LINE_ITEMS_TABLE_ID   = os.getenv("ORDER_LINE_ITEMS_TABLE")
ORDERS_TABLE_ID             = os.getenv("ORDERS_TABLE")
FRENCH_INVENTORIES_TABLE_ID = os.getenv("FRENCH_INVENTORIES_TABLE")

AMZ_CLIENT_ID     = os.getenv("CLIENT_ID")
AMZ_CLIENT_SECRET = os.getenv("CLIENT_SECRET")
AMZ_REFRESH_TOKEN = os.getenv("AMZ_REFRESH_TOKEN")
AWS_ACCESS_KEY    = os.getenv("AWS_ACCESS_KEY")
AWS_SECRET_KEY    = os.getenv("AWS_SECRET_KEY")
AWS_REGION        = os.getenv("AWS_REGION", "eu-west-1")
MARKETPLACE_ID    = "A2VIGQ35RCS4UG"  # UAE

AMZ_PRODUCTION = os.getenv("AMZ_PRODUCTION", "false").lower() == "true"
AMAZON_API_BASE = (
    "https://sellingpartnerapi-eu.amazon.com"
    if AMZ_PRODUCTION else
    "https://sandbox.sellingpartnerapi-eu.amazon.com"
)

def get_airtable_headers():
    return {
        "Authorization": f"Bearer {os.getenv('AIRTABLE_TOKEN')}",
        "Content-Type":  "application/json"
    }

aws_auth          = AWS4Auth(AWS_ACCESS_KEY, AWS_SECRET_KEY, AWS_REGION, "execute-api")
amazon_lock       = threading.Lock()
last_sync_time    = 0
MIN_SYNC_INTERVAL = 300  # minimum 5 minutes between syncs

# ── Startup log ────────────────────────────────────────────────────────────────
print("🚀 App starting...", flush=True)
print(f"🌍 Amazon mode: {'PRODUCTION' if AMZ_PRODUCTION else 'SANDBOX'}", flush=True)
print("AIRTABLE_TOKEN:",         bool(AIRTABLE_TOKEN), flush=True)
print("BASE_ID:",                bool(BASE_ID), flush=True)
print("CUSTOMERS_TABLE:",        bool(CUSTOMERS_TABLE_ID), flush=True)
print("ORDER_LINE_ITEMS:",       bool(ORDER_LINE_ITEMS_TABLE_ID), flush=True)
print("ORDERS_TABLE:",           bool(ORDERS_TABLE_ID), flush=True)
print("FRENCH_INVENTORIES:",     bool(FRENCH_INVENTORIES_TABLE_ID), flush=True)
print("CLIENT_ID:",              bool(AMZ_CLIENT_ID), flush=True)
print("AWS_ACCESS_KEY:",         bool(AWS_ACCESS_KEY), flush=True)


# ── Airtable helpers (Section 2) ───────────────────────────────────────────────

def airtable_search(table_id, formula):
    r = requests.get(
        f"{AIRTABLE_URL}/{BASE_ID}/{table_id}",
        headers=get_airtable_headers(),
        params={"filterByFormula": formula},
        timeout=REQUEST_TIMEOUT
    )
    r.raise_for_status()
    records = r.json().get("records", [])
    print(f"🔍 Found {len(records)} records", flush=True)
    return records

def airtable_create(table_id, fields):
    r = requests.post(
        f"{AIRTABLE_URL}/{BASE_ID}/{table_id}",
        headers=get_airtable_headers(),
        json={"fields": fields},
        timeout=REQUEST_TIMEOUT
    )
    if r.status_code >= 400:
        print("❌ Create error:", r.text, flush=True)
        r.raise_for_status()
    print("✅ Record created", flush=True)
    return r.json()

def airtable_update(table_id, record_id, fields):
    print(f"✏️ Updating {record_id}", flush=True)
    print(f"🧾 Fields being sent: {fields}", flush=True)
    r = requests.patch(
        f"{AIRTABLE_URL}/{BASE_ID}/{table_id}/{record_id}",
        headers=get_airtable_headers(),
        json={"fields": fields},
        timeout=REQUEST_TIMEOUT
    )
    print(f"🟡 Update status: {r.status_code}", flush=True)
    print(f"🟡 Update response: {r.text[:300]}", flush=True)
    if r.status_code >= 400:
        print("❌ Update error:", r.text, flush=True)
        r.raise_for_status()
    print("✅ Record updated", flush=True)


# ── Amazon helpers ─────────────────────────────────────────────────────────────

def get_amazon_token():
    print("🔑 Getting Amazon token...", flush=True)
    r = requests.post(
        "https://api.amazon.com/auth/o2/token",
        data={
            "grant_type":    "refresh_token",
            "refresh_token": AMZ_REFRESH_TOKEN,
            "client_id":     AMZ_CLIENT_ID,
            "client_secret": AMZ_CLIENT_SECRET,
        },
        timeout=REQUEST_TIMEOUT
    )
    r.raise_for_status()
    print("✅ Amazon token received", flush=True)
    return r.json()["access_token"]

def get_amazon_orders(token, days=2):
    print(f"📦 Fetching Amazon orders (last {days} days)...", flush=True)
    if AMZ_PRODUCTION:
        params = {
            "MarketplaceIds": MARKETPLACE_ID,
            "CreatedAfter":   (datetime.utcnow() - timedelta(days=days)).isoformat()
        }
    else:
        params = {
            "MarketplaceIds": "ATVPDKIKX0DER",
            "CreatedAfter":   "TEST_CASE_200"
        }
    r = requests.get(
        f"{AMAZON_API_BASE}/orders/v0/orders",
        headers={"x-amz-access-token": token, "Content-Type": "application/json"},
        params=params,
        auth=aws_auth,
        timeout=REQUEST_TIMEOUT
    )
    print("🟡 Orders status:", r.status_code, flush=True)
    print("🟡 Orders response:", r.text[:300], flush=True)
    r.raise_for_status()
    orders = r.json().get("payload", {}).get("Orders", [])
    print(f"✅ Amazon orders fetched: {len(orders)}", flush=True)
    return orders

def get_amazon_order_items(token, order_id):
    if not AMZ_PRODUCTION:
        print("🧪 Sandbox — using dummy item", flush=True)
        return [{
            "Title":           "Test Product",
            "SellerSKU":       "TEST-SKU-001",
            "QuantityOrdered": 1,
            "ItemPrice":       {"Amount": "99.99", "CurrencyCode": "USD"}
        }]
    print(f"📦 Fetching items for {order_id}", flush=True)
    r = requests.get(
        f"{AMAZON_API_BASE}/orders/v0/orders/{order_id}/orderItems",
        headers={"x-amz-access-token": token, "Content-Type": "application/json"},
        auth=aws_auth,
        timeout=REQUEST_TIMEOUT
    )
    print("🟡 Items status:", r.status_code, flush=True)
    r.raise_for_status()
    items = r.json().get("payload", {}).get("OrderItems", [])
    print(f"✅ Items found: {len(items)}", flush=True)
    return items

def get_rdt_token(access_token, order_id):
    print(f"🔐 Getting RDT for {order_id}", flush=True)
    r = requests.post(
        f"{AMAZON_API_BASE}/tokens/2021-03-01/restrictedDataToken",
        headers={
            "x-amz-access-token": access_token,
            "Content-Type":       "application/json"
        },
        json={
            "restrictedResources": [{
                "method":       "GET",
                "path":         f"/orders/v0/orders/{order_id}",
                "dataElements": ["buyerInfo"]
            }]
        },
        auth=aws_auth,
        timeout=REQUEST_TIMEOUT
    )
    print(f"🟡 RDT status: {r.status_code}", flush=True)
    if r.status_code == 200:
        return r.json().get("restrictedDataToken")
    print(f"⚠️ RDT failed: {r.text[:200]}", flush=True)
    return None

def get_order_with_pii(access_token, order_id):
    rdt = get_rdt_token(access_token, order_id)
    if not rdt:
        return {}
    r = requests.get(
        f"{AMAZON_API_BASE}/orders/v0/orders/{order_id}",
        headers={"x-amz-access-token": rdt, "Content-Type": "application/json"},
        auth=aws_auth,
        timeout=REQUEST_TIMEOUT
    )
    print(f"🟡 PII order status: {r.status_code}", flush=True)
    if r.status_code == 200:
        payload = r.json().get("payload", {})
        print(f"🔍 BuyerInfo: {payload.get('BuyerInfo', {})}", flush=True)
        print(f"🔍 ShippingAddress: {payload.get('ShippingAddress', {})}", flush=True)
        return payload
    print(f"⚠️ PII order failed: {r.text[:200]}", flush=True)
    return {}


# ── Status mappers (Section 2) ─────────────────────────────────────────────────

def map_shipping(status):
    s = status.lower()
    if s == "shipped":                return "Shipped"
    if s == "delivered":              return "Delivered"
    if s == "canceled":               return "Cancelled"
    if s in ["unshipped", "pending"]: return "New"
    return "New"

def map_payment(status):
    s = status.lower()
    if s in ["shipped", "delivered"]: return "Paid"
    if s == "canceled":               return "Failed"
    return "Pending"


# ── Customer helpers (Section 2) ───────────────────────────────────────────────

def get_or_create_customer(order, access_token=None):
    order_id    = order.get("AmazonOrderId", "")
    buyer_name  = ""
    buyer_email = ""
    buyer_phone = ""

    if AMZ_PRODUCTION and access_token:
        pii_order   = get_order_with_pii(access_token, order_id)
        buyer_info  = pii_order.get("BuyerInfo", {})
        buyer_email = buyer_info.get("BuyerEmail", "").strip()
        buyer_name  = buyer_info.get("BuyerName", "").strip()
        if not buyer_name:
            shipping    = pii_order.get("ShippingAddress", {})
            buyer_phone = shipping.get("Phone", "").strip()
            city        = shipping.get("City", "")
            country     = shipping.get("CountryCode", "")
            if city or country:
                buyer_name = f"Amazon Customer - {city}, {country}".strip(", ")
                print(f"📦 Using city/country: {buyer_name}", flush=True)
    else:
        buyer_info  = order.get("BuyerInfo", {})
        buyer_email = buyer_info.get("BuyerEmail", "").strip()
        buyer_name  = buyer_info.get("BuyerName", "").strip()

    if not buyer_name:
        buyer_name = "Amazon Customer"

    amazon_id = buyer_email if buyer_email else order_id
    print(f"👤 Amazon Id: {amazon_id} | name: {buyer_name}", flush=True)

    # Step 0: Search by Amazon Id first (prevents duplicates)
    records = airtable_search(CUSTOMERS_TABLE_ID, f"{{Amazon Id}}='{amazon_id}'")
    if records:
        print(f"👤 Found by Amazon Id", flush=True)
        return records[0]["id"]

    # Step 1: Search by email (Mail id)
    if buyer_email:
        records = airtable_search(CUSTOMERS_TABLE_ID, f"{{Mail id}}='{buyer_email}'")
        if records:
            print(f"👤 Found by email: {records[0]['fields'].get('Customer Name','')}", flush=True)
            airtable_update(CUSTOMERS_TABLE_ID, records[0]["id"], {"Amazon Id": amazon_id})
            return records[0]["id"]

    # Step 2: Search by phone (Contact Number)
    if buyer_phone:
        records = airtable_search(CUSTOMERS_TABLE_ID, f"{{Contact Number}}='{buyer_phone}'")
        if records:
            print(f"👤 Found by phone: {records[0]['fields'].get('Customer Name','')}", flush=True)
            airtable_update(CUSTOMERS_TABLE_ID, records[0]["id"], {"Amazon Id": amazon_id})
            return records[0]["id"]

    # Step 3: Create new customer
    print(f"👤 Creating new customer: {buyer_name}", flush=True)
    fields = {
        "Customer Name":          buyer_name,
        "Amazon Id":              amazon_id,
        "Acquired sales channel": "Amazon",
    }
    if buyer_email:
        fields["Mail id"] = buyer_email
    if buyer_phone:
        fields["Contact Number"] = buyer_phone
    result = airtable_create(CUSTOMERS_TABLE_ID, fields)
    return result["id"]


# ── Orders table helpers (Section 2) ──────────────────────────────────────────

def get_or_create_order(order_id, customer_id, order_date, pay, ship, ship_by=None):
    print(f"📋 Orders table lookup | {order_id}", flush=True)
    records = airtable_search(ORDERS_TABLE_ID, f"{{Order ID}}='{order_id}'")
    if records:
        existing_id   = records[0]["id"]
        print(f"📋 Existing order found — updating", flush=True)
        update_fields = {
            "Payment Status":  pay,
            "Shipping Status": ship,
        }
        if ship_by:
            update_fields["Ship By"] = ship_by
        airtable_update(ORDERS_TABLE_ID, existing_id, update_fields)
        return existing_id
    print(f"📋 Creating new order record", flush=True)
    fields = {
        "Order ID":        order_id,
        "Sales Channel":   "Amazon",
        "Order Date":      order_date,
        "Payment Status":  pay,
        "Shipping Status": ship,
    }
    if ship_by:
        fields["Ship By"] = ship_by
    if customer_id:
        fields["Customer"] = [customer_id]
    result = airtable_create(ORDERS_TABLE_ID, fields)
    return result["id"]


# ── French Inventories helpers (Section 2) ─────────────────────────────────────

def find_product_by_sku(sku):
    if not sku:
        print("⚠️ No SKU provided", flush=True)
        return None
    print(f"🔎 Looking up SKU: {sku}", flush=True)
    records = airtable_search(FRENCH_INVENTORIES_TABLE_ID, f"{{SKU}}='{sku}'")
    if records:
        print(f"✅ Product found for SKU: {sku}", flush=True)
        return records[0]["id"]
    print(f"⚠️ No product found for SKU: {sku}", flush=True)
    return None


# ── Order line items helpers (Section 2) ──────────────────────────────────────

def get_existing_line(order_id):
    records = airtable_search(
        ORDER_LINE_ITEMS_TABLE_ID,
        f"{{Order ID}}='{order_id}'"
    )
    return records[0]["id"] if records else None

def build_line_fields(order_id, product_title, order_date, qty, price,
                      pay, ship, customer_id, orders_record_id, product_record_id):
    fields = {
        "Order ID":            order_id,
        "Order Number":        order_id,
        "Amazon Product Name": product_title,
        "Order Date":          order_date,
        "Qty":                 qty,
        "Rate":                price,
        "Tax Type":            "5%",
        "Sales Channel":       "Amazon",
        "Payment Status":      pay,
        "Shipping Status":     ship,
    }
    if customer_id:
        fields["Customer"] = [customer_id]
    if orders_record_id:
        fields["Order"] = [orders_record_id]
    if product_record_id:
        fields["Product"] = [product_record_id]
    return fields


# ── Process one Amazon order ───────────────────────────────────────────────────

def process_order(order, token):
    order_id     = order.get("AmazonOrderId", "")
    order_status = order.get("OrderStatus", "")
    order_date   = order.get("PurchaseDate", "")[:10]
    pay          = map_payment(order_status)
    ship         = map_shipping(order_status)

    print(f"\n📦 Processing {order_id} | {order_status}", flush=True)

    # Extract Ship By date from Amazon order
    ship_by_raw = order.get("LatestShipDate", "") or order.get("EarliestShipDate", "")
    ship_by     = ship_by_raw[:10] if ship_by_raw else None
    if ship_by:
        print(f"📅 Ship By: {ship_by}", flush=True)

    # Step 1: Get or create customer
    customer_id = get_or_create_customer(order, token)

    # Step 2: Get or create order in Orders table
    orders_record_id = get_or_create_order(
        order_id, customer_id, order_date, pay, ship, ship_by
    )

    # Step 3: Get order items from Amazon
    try:
        items = get_amazon_order_items(token, order_id)
    except Exception as e:
        print(f"❌ Items fetch failed: {e}", flush=True)
        return

    # Step 4: Create/update Order Line Items
    for item in items:
        product_title     = item.get("Title", "")
        sku               = item.get("SellerSKU", "")
        qty               = int(item.get("QuantityOrdered", 1))
        price             = float(item.get("ItemPrice", {}).get("Amount", 0))
        product_record_id = find_product_by_sku(sku)
        existing_id       = get_existing_line(order_id)

        fields = build_line_fields(
            order_id, product_title, order_date, qty, price,
            pay, ship, customer_id, orders_record_id, product_record_id
        )

        if existing_id:
            airtable_update(ORDER_LINE_ITEMS_TABLE_ID, existing_id, fields)
            print(f"🔄 Updated line item for {order_id}", flush=True)
        else:
            airtable_create(ORDER_LINE_ITEMS_TABLE_ID, fields)
            print(f"✅ Created: {order_id} → {product_title}", flush=True)


# ── Main sync jobs (Section 2) ────────────────────────────────────────────────

def sync_amazon_orders_job():
    if not amazon_lock.acquire(blocking=False):
        print("⏳ Sync already running — skipped", flush=True)
        return

    print(f"⏰ Amazon sync started ({'PRODUCTION' if AMZ_PRODUCTION else 'SANDBOX'})", flush=True)

    try:
        token  = get_amazon_token()
        orders = get_amazon_orders(token, days=2)
        for order in orders:
            process_order(order, token)
    except Exception as e:
        print("❌ Sync error:", e, flush=True)
    finally:
        amazon_lock.release()
        print("🎉 Amazon sync finished", flush=True)


def sync_all_orders_job():
    if not amazon_lock.acquire(blocking=False):
        print("⏳ Sync already running — skipped", flush=True)
        return

    print("⏰ SYNC ALL started", flush=True)

    try:
        token         = get_amazon_token()
        all_orders    = []
        created_after = (datetime.utcnow() - timedelta(days=500)).isoformat()
        next_token    = None

        while True:
            if AMZ_PRODUCTION:
                params = {"MarketplaceIds": MARKETPLACE_ID, "CreatedAfter": created_after}
            else:
                params = {"MarketplaceIds": "ATVPDKIKX0DER", "CreatedAfter": "TEST_CASE_200"}

            if next_token:
                params["NextToken"] = next_token

            r = requests.get(
                f"{AMAZON_API_BASE}/orders/v0/orders",
                headers={"x-amz-access-token": token, "Content-Type": "application/json"},
                params=params,
                auth=aws_auth,
                timeout=REQUEST_TIMEOUT
            )
            r.raise_for_status()
            payload    = r.json().get("payload", {})
            orders     = payload.get("Orders", [])
            next_token = payload.get("NextToken")
            all_orders.extend(orders)
            print(f"📦 Fetched {len(orders)} | Total: {len(all_orders)}", flush=True)
            if not next_token:
                break

        print(f"✅ Total orders to sync: {len(all_orders)}", flush=True)
        for order in all_orders:
            process_order(order, token)

    except Exception as e:
        print("❌ Sync all error:", e, flush=True)
    finally:
        amazon_lock.release()
        print("🎉 SYNC ALL finished", flush=True)


def backfill_ship_by_job():
    print("🔄 Starting Amazon Ship By backfill...", flush=True)
    try:
        token         = get_amazon_token()
        all_orders    = []
        created_after = (datetime.utcnow() - timedelta(days=500)).isoformat()
        next_token    = None

        while True:
            params = {"MarketplaceIds": MARKETPLACE_ID, "CreatedAfter": created_after}
            if next_token:
                params["NextToken"] = next_token
            r = requests.get(
                f"{AMAZON_API_BASE}/orders/v0/orders",
                headers={"x-amz-access-token": token, "Content-Type": "application/json"},
                params=params,
                auth=aws_auth,
                timeout=REQUEST_TIMEOUT
            )
            r.raise_for_status()
            payload    = r.json().get("payload", {})
            orders     = payload.get("Orders", [])
            next_token = payload.get("NextToken")
            all_orders.extend(orders)
            print(f"📦 Fetched {len(orders)} | Total: {len(all_orders)}", flush=True)
            if not next_token:
                break

        print(f"✅ Total orders to backfill: {len(all_orders)}", flush=True)
        updated = 0
        skipped = 0

        for order in all_orders:
            order_id    = order.get("AmazonOrderId", "")
            ship_by_raw = order.get("LatestShipDate", "") or order.get("EarliestShipDate", "")
            ship_by     = ship_by_raw[:10] if ship_by_raw else None

            if not ship_by:
                skipped += 1
                continue

            records = airtable_search(ORDERS_TABLE_ID, f"{{Order ID}}='{order_id}'")
            if records:
                airtable_update(ORDERS_TABLE_ID, records[0]["id"], {"Ship By": ship_by})
                print(f"✅ {order_id} → Ship By: {ship_by}", flush=True)
                updated += 1
            else:
                print(f"⚠️ Not found in Airtable: {order_id}", flush=True)
                skipped += 1

        print(f"🎉 Backfill complete: {updated} updated, {skipped} skipped", flush=True)

    except Exception as e:
        print(f"❌ Backfill error: {e}", flush=True)


# ── Routes (Section 2) ────────────────────────────────────────────────────────

@app.route("/amazon", methods=["GET", "HEAD"])
# NOTE: renamed from / to /amazon to avoid conflict with Section 1
def amazon_health():
    return "OK", 200

@app.route("/wake", methods=["GET"])
def wake():
    return "awake", 200

@app.route("/ping", methods=["GET"])
def ping():
    print("🔥 /ping HIT", flush=True)
    received_secret = request.headers.get("X-Update-Secret")
    expected_secret = os.getenv("UPDATE_SECRET")
    if received_secret != expected_secret:
        print("⛔ Unauthorized", flush=True)
        return jsonify({"error": "Unauthorized"}), 401
    thread = threading.Thread(target=sync_amazon_orders_job)
    thread.daemon = True
    thread.start()
    return jsonify({
        "status": "Sync started",
        "mode":   "PRODUCTION" if AMZ_PRODUCTION else "SANDBOX"
    }), 200

@app.route("/sync-all", methods=["GET"])
def sync_all():
    print("🔥 /sync-all HIT", flush=True)
    thread = threading.Thread(target=sync_all_orders_job)
    thread.daemon = True
    thread.start()
    return jsonify({
        "status": "Full sync started — last 60 days",
        "mode":   "PRODUCTION" if AMZ_PRODUCTION else "SANDBOX"
    }), 200

@app.route("/backfill-ship-by", methods=["GET"])
def backfill_ship_by():
    print("🔥 /backfill-ship-by HIT", flush=True)
    thread = threading.Thread(target=backfill_ship_by_job)
    thread.daemon = True
    thread.start()
    return jsonify({
        "status":  "Backfill started — last 500 days",
        "message": "Watch Render logs for progress"
    }), 200

@app.route("/callback")
def callback():
    code = request.args.get("spapi_oauth_code")
    if not code:
        return jsonify({"error": "No code received", "args": dict(request.args)}), 400
    print(f"📥 OAuth code received: {code[:20]}...", flush=True)
    r = requests.post(
        "https://api.amazon.com/auth/o2/token",
        data={
            "grant_type":    "authorization_code",
            "code":          code,
            "client_id":     AMZ_CLIENT_ID,
            "client_secret": AMZ_CLIENT_SECRET,
        },
        timeout=REQUEST_TIMEOUT
    )
    print(f"🟡 Token exchange status: {r.status_code}", flush=True)
    if r.status_code != 200:
        return jsonify({"error": "Token exchange failed", "detail": r.json()}), 400
    refresh_token = r.json().get("refresh_token", "")
    print(f"✅ Refresh token received: {refresh_token[:30]}...", flush=True)
    return f"""
    <html><body style="font-family:monospace;padding:40px;background:#f0fff0">
    <h2 style="color:green">✅ Authorization successful!</h2>
    <p><b>Copy your Refresh Token and save it in Render as AMZ_REFRESH_TOKEN:</b></p>
    <div style="background:#fff;border:2px solid green;padding:20px;
                word-break:break-all;border-radius:8px;margin:20px 0">
        {refresh_token}
    </div>
    <p>Then set AMZ_PRODUCTION=true and redeploy.</p>
    </body></html>
    """, 200

@app.route("/download-orders", methods=["GET"])
def download_orders():
    print("🔥 /download-orders HIT", flush=True)
    try:
        token      = get_amazon_token()
        all_orders = []
        if AMZ_PRODUCTION:
            params = {
                "MarketplaceIds": MARKETPLACE_ID,
                "CreatedAfter":   (datetime.utcnow() - timedelta(days=500)).isoformat()
            }
        else:
            params = {"MarketplaceIds": "ATVPDKIKX0DER", "CreatedAfter": "TEST_CASE_200"}

        next_token = None
        while True:
            if next_token:
                params["NextToken"] = next_token
            r = requests.get(
                f"{AMAZON_API_BASE}/orders/v0/orders",
                headers={"x-amz-access-token": token, "Content-Type": "application/json"},
                params=params,
                auth=aws_auth,
                timeout=REQUEST_TIMEOUT
            )
            r.raise_for_status()
            payload    = r.json().get("payload", {})
            orders     = payload.get("Orders", [])
            next_token = payload.get("NextToken")
            all_orders.extend(orders)
            if not next_token:
                break

        print(f"✅ Total orders: {len(all_orders)}", flush=True)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Order ID", "Order Status", "Purchase Date",
            "Buyer Name", "Buyer Email", "Sales Channel",
            "Order Total", "Currency", "Fulfillment Channel",
            "Ship Service Level", "Product Name", "SKU",
            "Quantity", "Item Price",
        ])
        for order in all_orders:
            order_id      = order.get("AmazonOrderId", "")
            order_status  = order.get("OrderStatus", "")
            purchase_date = order.get("PurchaseDate", "")[:10]
            buyer_name    = order.get("BuyerInfo", {}).get("BuyerName", "")
            buyer_email   = order.get("BuyerInfo", {}).get("BuyerEmail", "")
            sales_channel = order.get("SalesChannel", "")
            order_total   = order.get("OrderTotal", {}).get("Amount", "")
            currency      = order.get("OrderTotal", {}).get("CurrencyCode", "")
            fulfillment   = order.get("FulfillmentChannel", "")
            ship_level    = order.get("ShipServiceLevel", "")
            try:
                items = get_amazon_order_items(token, order_id)
            except:
                items = []
            if items:
                for item in items:
                    writer.writerow([
                        order_id, order_status, purchase_date,
                        buyer_name, buyer_email, sales_channel,
                        order_total, currency, fulfillment, ship_level,
                        item.get("Title", ""), item.get("SellerSKU", ""),
                        item.get("QuantityOrdered", ""),
                        item.get("ItemPrice", {}).get("Amount", ""),
                    ])
            else:
                writer.writerow([
                    order_id, order_status, purchase_date,
                    buyer_name, buyer_email, sales_channel,
                    order_total, currency, fulfillment, ship_level,
                    "", "", "", ""
                ])
        output.seek(0)
        filename = f"amazon_orders_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print("❌ Download error:", e, flush=True)
        return jsonify({"error": str(e)}), 500

@app.route("/debug")
def debug():
    token = AIRTABLE_TOKEN or ""
    return jsonify({
        "token_length":             len(token),
        "token_start":              token[:10] if token else "EMPTY",
        "token_starts_with_pat":    token.startswith("pat"),
        "base_id":                  BASE_ID,
        "customers_table":          bool(CUSTOMERS_TABLE_ID),
        "order_line_items_table":   bool(ORDER_LINE_ITEMS_TABLE_ID),
        "orders_table":             bool(ORDERS_TABLE_ID),
        "french_inventories_table": bool(FRENCH_INVENTORIES_TABLE_ID),
        "amazon_mode":              "PRODUCTION" if AMZ_PRODUCTION else "SANDBOX",
    })

@app.route("/test-customer-search")
def test_customer_search():
    email = request.args.get("email", "")
    phone = request.args.get("phone", "")
    results = {}
    if email:
        records = airtable_search(CUSTOMERS_TABLE_ID, f"{{Mail id}}='{email}'")
        results["email_search"] = [r["fields"].get("Customer Name", "") for r in records]
    if phone:
        records = airtable_search(CUSTOMERS_TABLE_ID, f"{{Contact Number}}='{phone}'")
        results["phone_search"] = [r["fields"].get("Customer Name", "") for r in records]
    if not email and not phone:
        return jsonify({"error": "Provide ?email= or ?phone= parameter"}), 400
    return jsonify(results)

@app.route("/test-airtable-direct")
def test_airtable_direct():
    token = os.getenv("AIRTABLE_TOKEN")
    r = requests.get(
        f"https://api.airtable.com/v0/{BASE_ID}/{CUSTOMERS_TABLE_ID}",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        params={"maxRecords": 1}
    )
    return jsonify({
        "status":     r.status_code,
        "response":   r.json(),
        "token_used": token[:15] + "..."
    })

@app.route("/auto-sync", methods=["GET"])
def auto_sync():
    global last_sync_time
    now = time.time()
    if now - last_sync_time < MIN_SYNC_INTERVAL:
        remaining = int(MIN_SYNC_INTERVAL - (now - last_sync_time))
        print(f"⏳ Auto-sync skipped — wait {remaining}s", flush=True)
        return jsonify({"status": "skipped", "next_sync_in": remaining}), 200
    last_sync_time = now
    print("🔔 Auto-sync triggered by UptimeRobot", flush=True)
    thread = threading.Thread(target=sync_amazon_orders_job)
    thread.daemon = True
    thread.start()
    return jsonify({
        "status": "Sync started",
        "mode":   "PRODUCTION" if AMZ_PRODUCTION else "SANDBOX"
    }), 200

@app.route("/shopify-fulfillment", methods=["POST"])
def shopify_fulfillment():
    print("🛍️ Shopify fulfillment webhook received", flush=True)
    data = request.json
    if not data:
        return jsonify({"error": "No data"}), 400

    order_id     = str(data.get("id", ""))
    order_number = str(data.get("order_number", ""))
    fulfillments = data.get("fulfillments", [])

    if not fulfillments:
        print("⚠️ No fulfillments in payload", flush=True)
        return jsonify({"status": "no fulfillments"}), 200

    fulfilled_at   = fulfillments[0].get("created_at", "")
    fulfilled_date = fulfilled_at[:10] if fulfilled_at else datetime.utcnow().strftime("%Y-%m-%d")
    print(f"🛍️ Order {order_id} | #{order_number} fulfilled at {fulfilled_date}", flush=True)

    records = airtable_search(ORDERS_TABLE_ID, f"{{Order ID}}='{order_id}'")
    if not records:
        records = airtable_search(ORDERS_TABLE_ID, f"{{Order ID}}='{order_number}'")

    if records:
        airtable_update(ORDERS_TABLE_ID, records[0]["id"], {"Ship By": fulfilled_date})
        print(f"✅ Ship By updated to {fulfilled_date}", flush=True)
    else:
        print(f"⚠️ Order not found: {order_id} / #{order_number}", flush=True)

    return jsonify({"status": "ok"}), 200


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — TRENDYOL → AIRTABLE SYNC
# ══════════════════════════════════════════════════════════════════════════════

TRENDYOL_SELLER_ID = os.getenv("SELLER_ID")
TRENDYOL_API_KEY   = os.getenv("API_KEY")
TRENDYOL_API_SECRET = os.getenv("API_SECRET")

TRENDYOL_BASE_URL = "https://apigw.trendyol.com"

print("🔐 ENV CHECK:", flush=True)
print("AIRTABLE_TOKEN:", bool(AIRTABLE_TOKEN), flush=True)
print("BASE_ID:", bool(BASE_ID), flush=True)
print("CUSTOMERS_TABLE:", bool(CUSTOMERS_TABLE_ID), flush=True)
print("ORDERS_TABLE:", bool(ORDERS_TABLE_ID), flush=True)
print("ORDER_LINE_ITEMS_TABLE:", bool(ORDER_LINE_ITEMS_TABLE_ID), flush=True)
print("FRENCH_INVENTORIES_TABLE:", bool(FRENCH_INVENTORIES_TABLE_ID), flush=True)
print("SELLER_ID:", bool(TRENDYOL_SELLER_ID), flush=True)
print("API_KEY:", bool(TRENDYOL_API_KEY), flush=True)
print("API_SECRET:", bool(TRENDYOL_API_SECRET), flush=True)
print("--------------------------------------------------", flush=True)

AIRTABLE_HEADERS = {
    "Authorization": f"Bearer {AIRTABLE_TOKEN}",
    "Content-Type": "application/json"
}

basic_token = base64.b64encode(
    f"{TRENDYOL_API_KEY}:{TRENDYOL_API_SECRET}".encode()
).decode()

TRENDYOL_HEADERS = {
    "Authorization": f"Basic {basic_token}",
    "User-Agent": f"{TRENDYOL_SELLER_ID} - Self Integration",
    "Content-Type": "application/json",
    "storeFrontCode": "AE"
}

sync_lock = threading.Lock()


# ── Airtable helpers (Section 3)
# NOTE: renamed with ty_ prefix to avoid conflict with Section 2 helpers ──────

def ty_airtable_search(table_id, formula):
    print(f"🔍 Airtable search | table={table_id} | formula={formula}", flush=True)
    r = requests.get(
        f"{AIRTABLE_URL}/{BASE_ID}/{table_id}",
        headers=AIRTABLE_HEADERS,
        params={"filterByFormula": formula},
        timeout=REQUEST_TIMEOUT
    )
    r.raise_for_status()
    records = r.json().get("records", [])
    print(f"🔍 Found {len(records)} records", flush=True)
    return records

def ty_airtable_create(table_id, fields):
    print(f"📝 Creating Airtable record in table={table_id}", flush=True)
    print("🧾 Payload:", fields, flush=True)
    r = requests.post(
        f"{AIRTABLE_URL}/{BASE_ID}/{table_id}",
        headers=AIRTABLE_HEADERS,
        json={"fields": fields},
        timeout=REQUEST_TIMEOUT
    )
    if r.status_code >= 400:
        print("❌ Airtable error:", r.text, flush=True)
        r.raise_for_status()
    record = r.json()
    print("✅ Airtable record created:", record["id"], flush=True)
    return record["id"]

def ty_airtable_update(table_id, record_id, fields):
    print(f"✏️ Updating Airtable record {record_id} in table={table_id}", flush=True)
    print("🧾 Update payload:", fields, flush=True)
    r = requests.patch(
        f"{AIRTABLE_URL}/{BASE_ID}/{table_id}/{record_id}",
        headers=AIRTABLE_HEADERS,
        json={"fields": fields},
        timeout=REQUEST_TIMEOUT
    )
    if r.status_code >= 400:
        print("❌ Airtable update error:", r.text, flush=True)
        r.raise_for_status()
    print("✅ Airtable record updated", flush=True)


# ── Status mappers (Section 3) ─────────────────────────────────────────────────

def map_shipping_status(order):
    s = order.get("status", "").lower()
    if s == "delivered":
        return "Delivered"
    if s in ["shipped", "invoiced", "in_transit"]:
        return "In Transit"
    if s == "cancelled":
        return "Cancelled"
    return "New"

def map_payment_status(order):
    s = order.get("status", "").lower()
    if s in ["paid", "invoiced"]:
        return "Paid"
    if s == "cancelled":
        return "Failed"
    if s == "refunded":
        return "Refunded"
    return "Pending"


# ── Customer (Section 3)
# NOTE: renamed ty_get_or_create_customer to avoid conflict with Section 2 ─────

def ty_get_or_create_customer(c):
    print(f"👤 Processing customer {c['id']} | {c['name']}", flush=True)
    records = ty_airtable_search(
        CUSTOMERS_TABLE_ID,
        f"{{Trendyol Id}}='{c['id']}'"
    )
    if records:
        print("👤 Existing customer found", flush=True)
        return records[0]["id"]

    print("👤 Creating new customer", flush=True)
    record_id = ty_airtable_create(
        CUSTOMERS_TABLE_ID,
        {
            "Customer Name": c["name"],
            "Trendyol Id": c["id"]
        }
    )
    print("👤 Customer created:", record_id, flush=True)
    return record_id


# ── French Inventories — find product by SKU (Section 3) ──────────────────────

def get_french_inventory_record_id(merchant_sku):
    if not merchant_sku:
        print("⚠️ No merchantSku provided — skipping product lookup", flush=True)
        return None

    print(f"🔎 Looking up French Inventories | SKU={merchant_sku}", flush=True)
    records = ty_airtable_search(
        FRENCH_INVENTORIES_TABLE_ID,
        f"{{SKU}}='{merchant_sku}'"
    )
    if records:
        record_id = records[0]["id"]
        print(f"✅ Found French Inventory record: {record_id}", flush=True)
        return record_id

    print(f"⚠️ No French Inventory record found for SKU={merchant_sku}", flush=True)
    return None


# ── Orders table — get or create order (Section 3)
# NOTE: renamed ty_get_or_create_order to avoid conflict with Section 2 ────────

def ty_get_or_create_order(order_id, order_number, customer_id, order_date, pay, ship, ship_by=None):
    print(f"📋 Processing Orders table | Order ID={order_id}", flush=True)
    records = ty_airtable_search(
        ORDERS_TABLE_ID,
        f"{{Order ID}}='{order_id}'"
    )

    if records:
        existing_id = records[0]["id"]
        print(f"📋 Existing order found: {existing_id} — updating statuses", flush=True)
        update_fields = {
            "Payment Status": pay,
            "Shipping Status": ship
        }
        if ship_by:
            update_fields["Ship By"] = ship_by
        ty_airtable_update(ORDERS_TABLE_ID, existing_id, update_fields)
        return existing_id

    print(f"📋 Creating new order in Orders table", flush=True)
    create_fields = {
        "Order ID": order_id,
        "Customer": [customer_id],
        "Order Date": order_date,
        "Sales Channel": "Trendyol",
        "Payment Status": pay,
        "Shipping Status": ship
    }
    if ship_by:
        create_fields["Ship By"] = ship_by
    new_id = ty_airtable_create(ORDERS_TABLE_ID, create_fields)
    print(f"📋 Order created: {new_id}", flush=True)
    return new_id


# ── Order line items — duplicate check (Section 3) ────────────────────────────

def get_existing_order_line(order_id, product_name):
    print(f"🔁 Checking existing line | Order={order_id} | Product={product_name}", flush=True)
    records = ty_airtable_search(
        ORDER_LINE_ITEMS_TABLE_ID,
        f"AND({{Order ID}}='{order_id}', {{Trendyol Product Name}}='{product_name}')"
    )
    if records:
        record_id = records[0]["id"]
        print(f"🔁 Found existing record: {record_id}", flush=True)
        return record_id
    print("🔁 No existing record found", flush=True)
    return None


# ── Order line items — create (Section 3) ─────────────────────────────────────

def create_order_line(
    order_id, order_number, order_record_id,
    customer_id, date, pay, ship,
    product, qty, price,
    french_inventory_record_id
):
    print(f"🛒 Creating line item | {order_number} | {product}", flush=True)

    fields = {
        "Order ID": order_id,
        "Order Number": order_number,
        "Order Date": date,
        "Rate": price,
        "Qty": qty,
        "Trendyol Product Name": product,
        "Sales Channel": "Trendyol",
        "Payment Status": pay,
        "Shipping Status": ship,
        "Customer": [customer_id],
        "Order": [order_record_id],
    }

    if french_inventory_record_id:
        fields["Product"] = [french_inventory_record_id]

    ty_airtable_create(ORDER_LINE_ITEMS_TABLE_ID, fields)


# ── Order line items — update statuses (Section 3) ────────────────────────────

def update_order_line_statuses(record_id, pay, ship):
    print(f"🔄 Updating statuses for record {record_id} | Pay={pay} | Ship={ship}", flush=True)
    ty_airtable_update(
        ORDER_LINE_ITEMS_TABLE_ID,
        record_id,
        {
            "Payment Status": pay,
            "Shipping Status": ship
        }
    )


# ── Main sync logic (Section 3) ───────────────────────────────────────────────

def sync_trendyol_orders_job():
    if not sync_lock.acquire(blocking=False):
        print("⏳ Sync already running — skipped", flush=True)
        return

    print("⏰ Trendyol sync started", flush=True)

    try:
        r = requests.get(
            f"{TRENDYOL_BASE_URL}/integration/order/sellers/{TRENDYOL_SELLER_ID}/orders",
            headers=TRENDYOL_HEADERS,
            params={"page": 0, "size": 50},
            timeout=REQUEST_TIMEOUT
        )
        r.raise_for_status()

        orders = r.json().get("content", [])
        print(f"📦 Orders fetched: {len(orders)}", flush=True)

        for o in orders:
            print(f"\n{'='*50}", flush=True)
            print(f"📦 Processing order {o['orderNumber']}", flush=True)

            order_id     = str(o["id"])
            order_number = str(o["orderNumber"])

            order_date = datetime.utcfromtimestamp(
                o["orderDate"] / 1000
            ).strftime("%Y-%m-%d")

            pay  = map_payment_status(o)
            ship = map_shipping_status(o)

            # ── Extract Ship By from estimatedDeliveryStartDate ──
            ship_by = None
            est_ts = o.get("estimatedDeliveryStartDate")
            if est_ts:
                try:
                    ship_by = datetime.utcfromtimestamp(est_ts / 1000).strftime("%Y-%m-%d")
                except Exception:
                    ship_by = None
            print(f"📅 Ship By: {ship_by}", flush=True)

            # ── STEP 1: Get or create Customer ──────────────────
            customer_id = ty_get_or_create_customer({
                "id":   str(o["customerId"]),
                "name": f"{o.get('customerFirstName', '')} {o.get('customerLastName', '')}".strip()
            })

            # ── STEP 2: Get or create/update Order in Orders table ──
            order_record_id = ty_get_or_create_order(
                order_id, order_number,
                customer_id, order_date,
                pay, ship, ship_by
            )

            # ── STEP 3: Process each line item ──────────────────
            for line in o.get("lines", []):
                product      = line.get("productName", "")
                qty          = line.get("quantity", 1)
                price        = line.get("price", 0)
                merchant_sku = line.get("merchantSku", "")

                french_inventory_record_id = get_french_inventory_record_id(merchant_sku)
                existing_record_id = get_existing_order_line(order_id, product)

                if existing_record_id:
                    update_order_line_statuses(existing_record_id, pay, ship)
                    print(f"🔄 Updated statuses for {order_number} → {product}", flush=True)
                else:
                    create_order_line(
                        order_id, order_number, order_record_id,
                        customer_id, order_date, pay, ship,
                        product, qty, price,
                        french_inventory_record_id
                    )
                    print(f"✅ Created line item for {order_number} → {product}", flush=True)

    except Exception as e:
        print("❌ Sync error:", e, flush=True)

    finally:
        sync_lock.release()
        print("🎉 Trendyol sync finished", flush=True)


# ── Trendyol Ship By backfill (Section 3) ────────────────────────────────────

def backfill_trendyol_ship_by_job():
    print("🔄 Starting Trendyol Ship By backfill (all months)...", flush=True)
    try:
        updated = 0
        skipped = 0

        # Loop month by month from Jan 2026 to today
        from datetime import date
        start_month = date(2026, 1, 1)
        today       = datetime.utcnow().date()

        current = start_month
        while current <= today:
            # Build start/end timestamps for the month
            month_start = int(datetime(current.year, current.month, 1).timestamp() * 1000)
            if current.month == 12:
                next_month = date(current.year + 1, 1, 1)
            else:
                next_month = date(current.year, current.month + 1, 1)
            month_end = int(datetime(next_month.year, next_month.month, 1).timestamp() * 1000) - 1

            print(f"📅 Fetching {current.strftime('%Y-%m')}...", flush=True)
            page = 0
            while True:
                r = requests.get(
                    f"{TRENDYOL_BASE_URL}/integration/order/sellers/{TRENDYOL_SELLER_ID}/orders",
                    headers=TRENDYOL_HEADERS,
                    params={"page": page, "size": 50, "startDate": month_start, "endDate": month_end},
                    timeout=REQUEST_TIMEOUT
                )
                r.raise_for_status()
                data        = r.json()
                orders      = data.get("content", [])
                total_pages = data.get("totalPages", 1)
                print(f"  Page {page+1}/{total_pages} — {len(orders)} orders", flush=True)

                for o in orders:
                    order_id = str(o["id"])
                    est_ts   = o.get("estimatedDeliveryStartDate")
                    if not est_ts:
                        skipped += 1
                        continue
                    try:
                        ship_by = datetime.utcfromtimestamp(est_ts / 1000).strftime("%Y-%m-%d")
                    except Exception:
                        skipped += 1
                        continue

                    records = ty_airtable_search(ORDERS_TABLE_ID, f"{{Order ID}}='{order_id}'")
                    if records:
                        ty_airtable_update(ORDERS_TABLE_ID, records[0]["id"], {"Ship By": ship_by})
                        print(f"  ✅ {order_id} → Ship By: {ship_by}", flush=True)
                        updated += 1
                    else:
                        skipped += 1

                page += 1
                if page >= total_pages:
                    break

            current = next_month

        print(f"🎉 Trendyol backfill complete: {updated} updated, {skipped} skipped", flush=True)

    except Exception as e:
        print(f"❌ Trendyol backfill error: {e}", flush=True)


# ── Routes (Section 3) ────────────────────────────────────────────────────────

@app.route("/ping/trendyol", methods=["GET"])
# NOTE: renamed from /ping to /ping/trendyol to avoid conflict with Section 2
def trendyol_ping():
    print("🔥 /ping/trendyol endpoint HIT", flush=True)

    received_secret = request.headers.get("X-Update-Secret")
    expected_secret = os.getenv("UPDATE_SECRET")

    if received_secret != expected_secret:
        print("⛔ Unauthorized", flush=True)
        return jsonify({"error": "Unauthorized"}), 401

    print("🚀 Starting background sync", flush=True)
    thread = threading.Thread(target=sync_trendyol_orders_job)
    thread.daemon = True
    thread.start()

    return jsonify({"status": "Sync started in background"}), 200

@app.route("/backfill-ship-by/trendyol", methods=["GET"])
def backfill_trendyol_ship_by():
    print("🔥 /backfill-ship-by/trendyol HIT", flush=True)
    thread = threading.Thread(target=backfill_trendyol_ship_by_job)
    thread.daemon = True
    thread.start()
    return jsonify({
        "status":  "Trendyol Ship By backfill started",
        "message": "Watch Render logs for progress"
    }), 200

@app.route("/wake/trendyol", methods=["GET"])
# NOTE: renamed from /wake to /wake/trendyol to avoid conflict with Section 2
def trendyol_wake():
    print("🌅 Server woken up", flush=True)
    return "awake", 200

@app.route("/trendyol", methods=["GET", "HEAD"])
# NOTE: renamed from / to /trendyol to avoid conflict with Section 1
def trendyol_health():
    return "OK", 200


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — SHOPIFY ABANDONED CART → AIRTABLE
# ══════════════════════════════════════════════════════════════════════════════

AIRTABLE_BASE_ID    = os.environ.get("AIRTABLE_BASE_ID", BASE_ID)
SHOPIFY_SECRET      = os.environ.get("SHOPIFY_WEBHOOK_SECRET", "")
SHOPIFY_ADMIN_TOKEN = os.environ.get("SHOPIFY_ADMIN_TOKEN", "")

AT_BASE = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}"
AT_HEADERS = {
    "Authorization": f"Bearer {AIRTABLE_TOKEN}",
    "Content-Type":  "application/json",
}
SHOPIFY_HEADERS = {
    "X-Shopify-Access-Token": SHOPIFY_ADMIN_TOKEN,
    "Content-Type": "application/json",
}

TABLE_CUSTOMERS   = "Customers"
TABLE_INVENTORIES = "French Inventories"
TABLE_LEADS       = "Lead table"

print("=" * 60, flush=True)
print("[STARTUP] Shopify → Airtable Abandoned Cart Service", flush=True)
print(f"[STARTUP] Airtable Base ID : {AIRTABLE_BASE_ID}", flush=True)
print(f"[STARTUP] Customers table  : {TABLE_CUSTOMERS}", flush=True)
print(f"[STARTUP] Inventories table: {TABLE_INVENTORIES}", flush=True)
print(f"[STARTUP] Leads table      : {TABLE_LEADS}", flush=True)
print(f"[STARTUP] Webhook secret   : {'SET' if SHOPIFY_SECRET else 'NOT SET (verification skipped)'}", flush=True)
print(f"[STARTUP] Shopify store    : {SHOPIFY_STORE or 'NOT SET'}", flush=True)
print(f"[STARTUP] Shopify token    : {'SET' if SHOPIFY_ADMIN_TOKEN else 'NOT SET (sync disabled)'}", flush=True)
print("=" * 60, flush=True)


# ── Shopify webhook verification (Section 4) ──────────────────────────────────

def verify_webhook(raw_body: bytes, hmac_header: str) -> bool:
    if not SHOPIFY_SECRET:
        print("[WEBHOOK] No secret configured — skipping HMAC verification", flush=True)
        return True
    digest = hmac.new(SHOPIFY_SECRET.encode(), raw_body, hashlib.sha256).digest()
    computed = base64.b64encode(digest).decode()
    result = hmac.compare_digest(computed, hmac_header or "")
    print(f"[WEBHOOK] HMAC verification: {'PASSED' if result else 'FAILED'}", flush=True)
    return result


# ── Airtable helpers (Section 4) ──────────────────────────────────────────────

def at_get(table: str, formula: str) -> list:
    url = f"{AT_BASE}/{requests.utils.quote(table)}"
    print(f"[AIRTABLE GET] Table: {table}", flush=True)
    print(f"[AIRTABLE GET] Formula: {formula}", flush=True)
    resp = requests.get(url, headers=AT_HEADERS, params={"filterByFormula": formula})
    print(f"[AIRTABLE GET] Status: {resp.status_code}", flush=True)
    resp.raise_for_status()
    records = resp.json().get("records", [])
    print(f"[AIRTABLE GET] Records found: {len(records)}", flush=True)
    return records


def at_create(table: str, fields: dict) -> dict:
    url = f"{AT_BASE}/{requests.utils.quote(table)}"
    print(f"[AIRTABLE CREATE] Table: {table}", flush=True)
    print(f"[AIRTABLE CREATE] Fields: {fields}", flush=True)
    resp = requests.post(url, headers=AT_HEADERS, json={"fields": fields})
    print(f"[AIRTABLE CREATE] Status: {resp.status_code}", flush=True)
    if not resp.ok:
        print(f"[AIRTABLE CREATE] Error: {resp.text}", flush=True)
    resp.raise_for_status()
    record = resp.json()
    print(f"[AIRTABLE CREATE] Created record ID: {record.get('id')}", flush=True)
    return record


# ── Business logic (Section 4) ────────────────────────────────────────────────

def find_customer(phone: str, email: str) -> dict | None:
    """
    Name/Mobile/Mail is a FORMULA field (read-only):
      CONCATENATE({Customer Name},"/",{Contact Number},"/",{Mail id})
    Search by actual underlying fields: Contact Number and Mail id.
    """
    print(f"\n[CUSTOMER SEARCH] phone={phone!r}  email={email!r}", flush=True)
    parts = []
    if phone:
        stripped = ''.join(c for c in phone if c.isdigit())
        if stripped:
            parts.append(f"{{Whatsapp number}}='{stripped}'")
    if email:
        parts.append(f"LOWER({{Mail id}})='{email.lower()}'")

    if not parts:
        print("[CUSTOMER SEARCH] No contact info — cannot search", flush=True)
        return None

    formula = f"OR({','.join(parts)})"
    records = at_get(TABLE_CUSTOMERS, formula)

    if records:
        rec = records[0]
        print(f"[CUSTOMER SEARCH] Found: ID={rec['id']}  Name={rec.get('fields',{}).get('Customer Name')}", flush=True)
        return rec
    print("[CUSTOMER SEARCH] Not found", flush=True)
    return None


def create_customer(name: str, phone: str, email: str) -> dict:
    """
    Write only editable fields. Name/Mobile/Mail auto-computes.
    """
    print(f"\n[CUSTOMER CREATE] name={name!r}  phone={phone!r}  email={email!r}", flush=True)
    fields: dict = {}
    if name:  fields["Customer Name"]  = name
    if phone: fields["Whatsapp number"] = phone
    if email: fields["Mail id"]        = email
    record = at_create(TABLE_CUSTOMERS, fields)
    print(f"[CUSTOMER CREATE] New customer ID: {record.get('id')}", flush=True)
    return record


def cart_find_product_by_sku(sku: str) -> dict | None:
    # NOTE: renamed from find_product_by_sku to cart_find_product_by_sku
    # to avoid conflict with Section 2's find_product_by_sku
    print(f"\n[PRODUCT SEARCH] SKU={sku!r}", flush=True)
    if not sku:
        print("[PRODUCT SEARCH] Empty SKU — skip", flush=True)
        return None
    records = at_get(TABLE_INVENTORIES, f"{{SKU}}='{sku}'")
    if records:
        rec = records[0]
        print(f"[PRODUCT SEARCH] Found: ID={rec['id']}  Name={rec.get('fields',{}).get('Product Name','?')}", flush=True)
        return rec
    print(f"[PRODUCT SEARCH] SKU '{sku}' NOT found in {TABLE_INVENTORIES}", flush=True)
    return None


def lead_exists_for_customer(customer_id: str) -> bool:
    """Check if a lead already exists for this customer to avoid duplicates."""
    formula = f"FIND('{customer_id}', ARRAYJOIN(Customers, ','))"
    records = at_get(TABLE_LEADS, formula)
    return len(records) > 0


def create_lead(customer_id: str, product_ids: list[str], abandoned_date: str) -> dict:
    print(f"\n[LEAD CREATE] customer_id={customer_id}", flush=True)
    print(f"[LEAD CREATE] product_ids={product_ids}", flush=True)
    print(f"[LEAD CREATE] date={abandoned_date}", flush=True)
    fields = {
        "Customers":           [customer_id],
        "Interested products": product_ids,
        "Lead created date":   abandoned_date,
        "Lead Source":         "Abandoned cart",
    }
    record = at_create(TABLE_LEADS, fields)
    print(f"[LEAD CREATE] Lead ID: {record.get('id')}", flush=True)
    return record


def process_single_checkout(checkout: dict) -> dict:
    """
    Shared logic used by both the webhook and the sync route.
    Returns a result dict describing what happened.
    """
    checkout_id = checkout.get("id")

    # Skip completed checkouts
    if checkout.get("completed_at"):
        print(f"[PROCESS] Checkout {checkout_id} already completed — skipping", flush=True)
        return {"status": "skipped", "reason": "already completed", "checkout_id": checkout_id}

    # Extract contact info
    cust    = checkout.get("customer") or {}
    billing = checkout.get("billing_address") or {}
    first = (cust.get("first_name") or billing.get("first_name") or "").strip()
    last  = (cust.get("last_name")  or billing.get("last_name")  or "").strip()
    name  = f"{first} {last}".strip() or billing.get("name", "Unknown")
    email = (cust.get("email") or checkout.get("email") or "").strip().lower()
    phone = (cust.get("phone") or billing.get("phone") or checkout.get("phone") or "").strip()

    print(f"[EXTRACT] Name : {name!r}", flush=True)
    print(f"[EXTRACT] Email: {email!r}", flush=True)
    print(f"[EXTRACT] Phone: {phone!r}", flush=True)

    if not email and not phone:
        print("[EXTRACT] No contact info — skipping", flush=True)
        return {"status": "skipped", "reason": "no contact info", "checkout_id": checkout_id}

    # Parse date
    raw_date = checkout.get("created_at", "")
    try:
        abandoned_date = datetime.fromisoformat(raw_date.replace("Z", "+00:00")).strftime("%Y-%m-%d")
    except Exception:
        abandoned_date = datetime.utcnow().strftime("%Y-%m-%d")
    print(f"[EXTRACT] Abandoned date: {abandoned_date}", flush=True)

    # Line items
    line_items = checkout.get("line_items", [])
    print(f"[EXTRACT] Line items ({len(line_items)}):", flush=True)
    for i, item in enumerate(line_items, 1):
        print(f"  [{i}] title={item.get('title')!r}  sku={item.get('sku')!r}  qty={item.get('quantity')}", flush=True)

    # STEP 1 — Find or create customer
    print("\n[STEP 1] Customer lookup...", flush=True)
    customer_record = find_customer(phone, email)
    customer_action = "found"
    if customer_record:
        print(f"[STEP 1] Existing customer: {customer_record['id']}", flush=True)
    else:
        print("[STEP 1] Not found — creating new customer", flush=True)
        customer_record = create_customer(name, phone, email)
        customer_action = "created"
        print(f"[STEP 1] New customer: {customer_record['id']}", flush=True)
    customer_id = customer_record["id"]

    # STEP 2 — Match SKUs
    print("\n[STEP 2] SKU matching...", flush=True)
    product_ids: list[str] = []
    unmatched_skus: list[str] = []
    for item in line_items:
        sku = (item.get("sku") or "").strip()
        if not sku:
            print(f"  [SKIP] '{item.get('title')}' has no SKU", flush=True)
            continue
        prod = cart_find_product_by_sku(sku)
        if prod:
            product_ids.append(prod["id"])
            print(f"  [OK] {sku} -> {prod['id']}", flush=True)
        else:
            unmatched_skus.append(sku)
            print(f"  [MISS] {sku} not found", flush=True)
    print(f"[STEP 2] Matched={len(product_ids)}  Unmatched={unmatched_skus}", flush=True)

    if not product_ids:
        print("[STEP 2] No matched products — lead NOT created", flush=True)
        return {
            "status":         "skipped",
            "reason":         "no matching SKUs",
            "checkout_id":    checkout_id,
            "customer_id":    customer_id,
            "customer_action": customer_action,
            "unmatched_skus": unmatched_skus,
        }

    # STEP 3 — Create lead
    print("\n[STEP 3] Creating lead...", flush=True)
    lead = create_lead(customer_id, product_ids, abandoned_date)
    lead_id = lead.get("id")
    print(f"\n[DONE] customer_id={customer_id}  lead_id={lead_id}  products={len(product_ids)}  unmatched={unmatched_skus}", flush=True)

    return {
        "status":          "success",
        "checkout_id":     checkout_id,
        "customer_id":     customer_id,
        "customer_action": customer_action,
        "lead_id":         lead_id,
        "products_linked": len(product_ids),
        "unmatched_skus":  unmatched_skus,
    }


# ── Webhook route (Section 4) ─────────────────────────────────────────────────

@app.route("/webhook/abandoned-checkout", methods=["POST"])
def abandoned_checkout():
    print("\n" + "=" * 60, flush=True)
    print(f"[WEBHOOK] Received at {datetime.utcnow().isoformat()}Z", flush=True)

    if not verify_webhook(request.data, request.headers.get("X-Shopify-Hmac-SHA256", "")):
        print("[WEBHOOK] Rejected — HMAC mismatch", flush=True)
        return jsonify({"error": "Unauthorized"}), 401

    checkout = request.get_json(force=True)
    if not checkout:
        print("[WEBHOOK] No JSON payload", flush=True)
        return jsonify({"error": "No payload"}), 400

    print(f"[WEBHOOK] Checkout ID   : {checkout.get('id')}", flush=True)
    print(f"[WEBHOOK] Checkout token: {checkout.get('token', 'N/A')}", flush=True)

    result = process_single_checkout(checkout)
    print("=" * 60, flush=True)

    if result["status"] == "skipped":
        return jsonify(result), 200
    return jsonify(result), 200


# ── Global sync state (Section 4) ─────────────────────────────────────────────
sync_state = {
    "running": False,
    "started_at": None,
    "stats": {},
    "last_error": None,
}


def run_sync_in_background(max_limit, since_date):
    """Runs the full Shopify → Airtable sync in a background thread."""
    global sync_state
    sync_state["running"]    = True
    sync_state["started_at"] = datetime.utcnow().isoformat() + "Z"
    sync_state["last_error"] = None
    stats = sync_state["stats"] = {
        "fetched": 0, "success": 0,
        "skipped_completed": 0, "skipped_no_contact": 0,
        "skipped_no_sku": 0, "duplicate_lead": 0, "errors": 0,
    }

    try:
        url    = f"https://{SHOPIFY_STORE}/admin/api/2024-04/checkouts.json"
        params = {"limit": 250, "status": "open"}
        if since_date:
            params["created_at_min"] = since_date

        page = 1
        done = False

        while url and not done:
            print(f"\n[SYNC] Fetching Shopify page {page}...", flush=True)
            resp = requests.get(url, headers=SHOPIFY_HEADERS, params=params)

            if resp.status_code == 429:
                wait = int(resp.headers.get("Retry-After", 2))
                print(f"[SYNC] Rate limited — waiting {wait}s", flush=True)
                time.sleep(wait)
                continue

            if not resp.ok:
                msg = f"Shopify API error: {resp.status_code} {resp.text}"
                print(f"[SYNC] {msg}", flush=True)
                sync_state["last_error"] = msg
                break

            checkouts = resp.json().get("checkouts", [])
            print(f"[SYNC] Page {page}: {len(checkouts)} checkouts received", flush=True)

            for checkout in checkouts:
                stats["fetched"] += 1
                print(f"\n[SYNC] ── Checkout #{checkout.get('id')} ({stats['fetched']}) ──", flush=True)

                cust  = checkout.get("customer") or {}
                phone = (cust.get("phone") or checkout.get("phone") or "").strip()
                email = (cust.get("email") or checkout.get("email") or "").strip().lower()
                try:
                    existing_customer = find_customer(phone, email) if (phone or email) else None
                except Exception:
                    existing_customer = None
                if existing_customer and lead_exists_for_customer(existing_customer["id"]):
                    print(f"[SYNC] Lead already exists for customer {existing_customer['id']} — skipping", flush=True)
                    stats["duplicate_lead"] += 1
                    continue

                try:
                    result = process_single_checkout(checkout)
                    if result["status"] == "success":
                        stats["success"] += 1
                    elif result.get("reason") == "already completed":
                        stats["skipped_completed"] += 1
                    elif result.get("reason") == "no contact info":
                        stats["skipped_no_contact"] += 1
                    elif result.get("reason") == "no matching SKUs":
                        stats["skipped_no_sku"] += 1
                except Exception as e:
                    print(f"[SYNC] ERROR on checkout {checkout.get('id')}: {e}", flush=True)
                    stats["errors"] += 1

                time.sleep(0.3)

                if max_limit and stats["fetched"] >= max_limit:
                    print(f"[SYNC] Reached limit of {max_limit} — stopping", flush=True)
                    done = True
                    break

            # Pagination
            link = resp.headers.get("Link", "")
            url  = None
            params = {}
            if 'rel="next"' in link:
                for part in link.split(","):
                    if 'rel="next"' in part:
                        url = part.split(";")[0].strip().strip("<>")
                        break
            page += 1

    except Exception as e:
        print(f"[SYNC] Fatal error: {e}", flush=True)
        sync_state["last_error"] = str(e)
    finally:
        sync_state["running"] = False
        print(f"\n[SYNC] Background thread complete — {stats}", flush=True)
        print("=" * 60, flush=True)


@app.route("/sync/abandoned-checkouts", methods=["GET", "POST"])
def sync_abandoned_checkouts():
    print("\n" + "=" * 60, flush=True)
    print(f"[SYNC] Request received at {datetime.utcnow().isoformat()}Z", flush=True)

    if not SHOPIFY_STORE or not SHOPIFY_ADMIN_TOKEN:
        print("[SYNC] SHOPIFY_STORE or SHOPIFY_ADMIN_TOKEN not set", flush=True)
        return jsonify({"error": "SHOPIFY_STORE and SHOPIFY_ADMIN_TOKEN env vars required"}), 500

    # Check status only — don't start a new sync if one is running
    if request.args.get("status") == "1":
        return jsonify({"sync_state": sync_state}), 200

    if sync_state["running"]:
        print("[SYNC] Already running — returning current progress", flush=True)
        return jsonify({
            "message":    "Sync already in progress",
            "sync_state": sync_state,
        }), 200

    body       = request.get_json(force=True, silent=True) or {}
    max_limit  = body.get("limit", None)
    since_date = body.get("since", None)
    print(f"[SYNC] Starting background thread: max_limit={max_limit}  since_date={since_date}", flush=True)

    thread = threading.Thread(
        target=run_sync_in_background,
        args=(max_limit, since_date),
        daemon=True,
    )
    thread.start()

    return jsonify({
        "message":    "Sync started in background. Check Render logs for progress.",
        "status_url": "/sync/abandoned-checkouts?status=1",
        "params":     {"limit": max_limit, "since": since_date},
    }), 200


# ── Health check (Section 4) ──────────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    print("[HEALTH] Health check", flush=True)
    return jsonify({"status": "ok", "service": "shopify-airtable-abandoned-cart"}), 200


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — DAILY SESSION REPORT → SLACK (#session)
# ══════════════════════════════════════════════════════════════════════════════
# NOTE: reuses SHOPIFY_STORE / SHOPIFY_ACCESS_TOKEN already defined in Section 1.
# Falls back to SHOPIFY_ADMIN_API_TOKEN if that's the env var name configured
# for this particular service on Render.

SESSION_REPORT_SHOPIFY_TOKEN = os.getenv("SHOPIFY_ADMIN_API_TOKEN") or SHOPIFY_ACCESS_TOKEN
SLACK_BOT_TOKEN               = os.getenv("SLACK_BOT_TOKEN")        # xoxb-...
SESSION_REPORT_CHANNEL_ID     = os.getenv("SLACK_CHANNEL_ID", "C0B9V9U312L")

IST = ZoneInfo("Asia/Kolkata")


# ── SHOPIFY: fetch sessions by landing page (Section 5) ───────────────────────

def fetch_sessions(date_str):
    print(f"[fetch_sessions] Querying Shopify for date: {date_str}", flush=True)

    url = f"https://{SHOPIFY_STORE}/admin/api/2026-04/graphql.json"
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": SESSION_REPORT_SHOPIFY_TOKEN,
    }

    query = f"""
    {{
      shopifyqlQuery(query: "FROM sessions SHOW landing_page_type, landing_page_path, online_store_visitors, sessions SINCE {date_str} UNTIL {date_str} GROUP BY landing_page_type, landing_page_path ORDER BY sessions DESC") {{
        tableData {{
          columns {{ name }}
          rows
        }}
        parseErrors
      }}
    }}
    """

    print(f"[fetch_sessions] Sending ShopifyQL request...", flush=True)
    resp = requests.post(url, json={"query": query}, headers=headers, timeout=30)
    print(f"[fetch_sessions] Response status: {resp.status_code}", flush=True)
    data = resp.json()

    if "errors" in data:
        print(f"[fetch_sessions] GraphQL errors: {data['errors']}", flush=True)
        raise RuntimeError(f"GraphQL errors: {data['errors']}")

    shopify_data = data.get("data", {}).get("shopifyqlQuery", {})
    parse_errors = shopify_data.get("parseErrors")
    if parse_errors:
        print(f"[fetch_sessions] ShopifyQL parse errors: {parse_errors}", flush=True)
        raise ValueError(f"ShopifyQL parse errors: {parse_errors}")

    table = shopify_data.get("tableData", {})
    if not table:
        print(f"[fetch_sessions] WARNING - No tableData returned.", flush=True)
        return []

    columns = [col["name"] for col in table.get("columns", [])]
    rows    = table.get("rows", [])
    print(f"[fetch_sessions] Columns: {columns}", flush=True)
    print(f"[fetch_sessions] Total rows: {len(rows)}", flush=True)

    results = []
    for row in rows:
        if isinstance(row, list):
            record = dict(zip(columns, row))
        elif isinstance(row, dict):
            record = row
        else:
            continue
        results.append(record)

    print(f"[fetch_sessions] Successfully parsed {len(results)} rows.", flush=True)
    return results


# ── BUILD EXCEL (Section 5) ────────────────────────────────────────────────────

def build_excel(rows, date_str):
    """
    Filters Product rows, keeps 3 columns, returns Excel bytes.
    """
    print(f"[build_excel] Starting Excel build for {date_str}. Input rows: {len(rows)}", flush=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessions by Landing Page"

    # Header
    headers = ["Landing page path", "Online store visitors", "Sessions"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows — only Product type
    count = 0
    skipped = 0
    for r in rows:
        page_type = str(r.get("landing_page_type", "")).strip().lower()
        if page_type != "product":
            print(f"[build_excel] Skipping row — type='{page_type}', path='{r.get('landing_page_path', '')}'", flush=True)
            skipped += 1
            continue
        ws.append([
            r.get("landing_page_path", ""),
            r.get("online_store_visitors", 0),
            r.get("sessions", 0),
        ])
        count += 1

    print(f"[build_excel] Rows added (Product): {count} | Rows skipped (non-Product): {skipped}", flush=True)

    # Column widths
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_size_kb = round(len(buf.getvalue()) / 1024, 2)
    print(f"[build_excel] Excel file built successfully. Size: {excel_size_kb} KB", flush=True)
    return buf.getvalue(), count


# ── SEND TO SLACK (Section 5) ─────────────────────────────────────────────────

def send_session_report_to_slack(excel_bytes, date_str, row_count):
    # NOTE: renamed from send_to_slack to send_session_report_to_slack
    # to keep this service's Slack upload flow distinct from other sections.
    filename = f"sessions_by_landing_page_{date_str}.xlsx"
    file_size = len(excel_bytes)
    print(f"[send_to_slack] Preparing to upload '{filename}' to Slack channel {SESSION_REPORT_CHANNEL_ID}", flush=True)
    print(f"[send_to_slack] File size: {round(file_size/1024, 2)} KB | Product rows: {row_count}", flush=True)

    # Step 1 — Get upload URL
    print(f"[send_to_slack] Step 1: Requesting upload URL from Slack...", flush=True)
    url_resp = requests.post(
        "https://slack.com/api/files.getUploadURLExternal",
        headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"},
        data={"filename": filename, "length": file_size},
        timeout=30,
    )
    url_resp.raise_for_status()
    url_result = url_resp.json()
    print(f"[send_to_slack] Step 1 response: ok={url_result.get('ok')} error={url_result.get('error','none')}", flush=True)

    if not url_result.get("ok"):
        raise RuntimeError(f"Slack getUploadURLExternal failed: {url_result.get('error')}")

    upload_url = url_result["upload_url"]
    file_id    = url_result["file_id"]
    print(f"[send_to_slack] Got upload URL. File ID: {file_id}", flush=True)

    # Step 2 — Upload file bytes to the upload URL
    print(f"[send_to_slack] Step 2: Uploading file bytes...", flush=True)
    upload_resp = requests.post(
        upload_url,
        files={"file": (filename, excel_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=60,
    )
    upload_resp.raise_for_status()
    print(f"[send_to_slack] Step 2 done. HTTP status: {upload_resp.status_code}", flush=True)

    # Step 3 — Complete upload and share to channel
    print(f"[send_to_slack] Step 3: Completing upload and posting to channel...", flush=True)
    complete_resp = requests.post(
        "https://slack.com/api/files.completeUploadExternal",
        headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"},
        json={
            "files": [{"id": file_id, "title": filename}],
            "channel_id": SESSION_REPORT_CHANNEL_ID,
            "initial_comment": (
                f":bar_chart: *Sessions by Landing Page — {date_str}*\n"
                f"Product pages: *{row_count} rows* | Columns: Landing page path, Online store visitors, Sessions"
            ),
        },
        timeout=30,
    )
    complete_resp.raise_for_status()
    complete_result = complete_resp.json()
    print(f"[send_to_slack] Step 3 response: ok={complete_result.get('ok')} error={complete_result.get('error','none')}", flush=True)

    if not complete_result.get("ok"):
        raise RuntimeError(f"Slack completeUploadExternal failed: {complete_result.get('error')}")

    print(f"[send_to_slack] SUCCESS - File posted to Slack channel. File ID: {file_id}", flush=True)
    print(f"[session-slack] Sent {row_count} product rows for {date_str} to Slack.", flush=True)


# ── MAIN JOB (Section 5) ───────────────────────────────────────────────────────

def run_session_report_job():
    # NOTE: renamed from run_job to run_session_report_job to avoid ambiguity
    # alongside the other background jobs defined in earlier sections.
    print(f"[run_job] ─────────────────────────────────────────", flush=True)
    print(f"[run_job] Job started at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    try:
        yesterday = (datetime.now(IST) - timedelta(days=1)).strftime("%Y-%m-%d")
        print(f"[run_job] Target date (yesterday): {yesterday}", flush=True)

        rows = fetch_sessions(yesterday)
        print(f"[run_job] Got {len(rows)} total rows from Shopify.", flush=True)

        excel_bytes, count = build_excel(rows, yesterday)
        print(f"[run_job] Filtered to {count} Product rows.", flush=True)

        send_session_report_to_slack(excel_bytes, yesterday, count)
        print(f"[run_job] Job completed successfully.", flush=True)

    except Exception as e:
        print(f"[run_job] ERROR: {e}", flush=True)
    print(f"[run_job] ─────────────────────────────────────────", flush=True)


# ── Routes (Section 5) ─────────────────────────────────────────────────────────

@app.route("/session-report/health", methods=["GET"])
# NOTE: renamed from /health to /session-report/health to avoid conflict with Section 4
def session_report_health():
    print(f"[session-report/health] Health check called at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    return jsonify({"status": "ok", "service": "session-report-slack"})


@app.route("/session-report/run-now", methods=["GET"])
# NOTE: renamed from /run-now to /session-report/run-now for a unique, namespaced path
def session_report_run_now():
    """Trigger the job manually for testing."""
    print(f"[session-report/run-now] Manual trigger called at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    run_session_report_job()
    return jsonify({"status": "done"})


@app.route("/session-report/run-date/<date_str>", methods=["GET"])
# NOTE: renamed from /run-date/<date_str> to /session-report/run-date/<date_str>
def session_report_run_date(date_str):
    """Fetch a specific date. Format: YYYY-MM-DD  e.g. /session-report/run-date/2026-06-07"""
    print(f"[session-report/run-date] Manual date trigger called for: {date_str}", flush=True)
    try:
        rows = fetch_sessions(date_str)
        excel_bytes, count = build_excel(rows, date_str)
        send_session_report_to_slack(excel_bytes, date_str, count)
        print(f"[session-report/run-date] Completed for {date_str}. Product rows: {count}", flush=True)
        return jsonify({"status": "done", "product_rows": count, "date": date_str})
    except Exception as e:
        print(f"[session-report/run-date] ERROR for {date_str}: {e}", flush=True)
        return jsonify({"status": "error", "message": str(e)}), 500


# ── Scheduler (Section 5) ──────────────────────────────────────────────────────
session_report_scheduler = BackgroundScheduler(timezone=IST)
session_report_scheduler.add_job(run_session_report_job, "cron", hour=17, minute=0)
session_report_scheduler.start()
print(f"[startup] Session report scheduler started. Job will run daily at 17:00 IST.", flush=True)
print(f"[startup] SESSION REPORT SHOPIFY_STORE     : {SHOPIFY_STORE}", flush=True)
print(f"[startup] SESSION REPORT SHOPIFY_TOKEN set : {'YES' if SESSION_REPORT_SHOPIFY_TOKEN else 'NO ⚠️'}", flush=True)
print(f"[startup] SESSION REPORT SLACK_TOKEN set   : {'YES' if SLACK_BOT_TOKEN else 'NO ⚠️'}", flush=True)
print(f"[startup] SESSION REPORT SLACK_CHANNEL_ID  : {SESSION_REPORT_CHANNEL_ID}", flush=True)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — DANA BOOKS → AIRTABLE COST SYNC (French Inventories)
# ══════════════════════════════════════════════════════════════════════════════
# NOTE: reuses AIRTABLE_BASE_ID already defined in Section 4 instead of
# redeclaring it. Reuses AIRTABLE_TOKEN as a fallback if AIRTABLE_API_KEY
# isn't set separately for this service.

DANABOOKS_URL        = "https://transactionhub.zerobook.shop/api/v1/transaction-history"
DANABOOKS_TOKEN      = os.environ.get("DANABOOKS_TOKEN", "")
DANABOOKS_IDENTIFIER = os.environ.get("DANABOOKS_IDENTIFIER", "thirdparty@danabooks.com")

AIRTABLE_API_KEY     = os.environ.get("AIRTABLE_API_KEY") or AIRTABLE_TOKEN
DANABOOKS_TABLE_NAME = "French Inventories"

DANABOOKS_IST = pytz.timezone("Asia/Kolkata")
# NOTE: kept separate from Section 5's `IST` (zoneinfo.ZoneInfo) — same
# timezone, different library object, so it's not safe to just reuse that name.

# Number of SKUs per Dana Books batch request
DANABOOKS_BATCH_SIZE  = 500
# Delay between batch requests (seconds)
DANABOOKS_BATCH_DELAY = 1.0
# Max retries on 429/520
DANABOOKS_MAX_RETRIES = 3
# Wait time before retry (seconds)
DANABOOKS_RETRY_WAIT  = 10

# Thread-safe counters
_danabooks_progress_lock = threading.Lock()


# ── Dana Books helpers (Section 6) ─────────────────────────────────────────────

def get_prices_for_batch(skus):
    """
    Fetch latest purchase prices from Dana Books for a batch of SKUs.
    Returns dict: { sku: float_price_or_None, ... }
    """
    headers = {
        "Authorization": f"Bearer {DANABOOKS_TOKEN}",
        "Identifier": DANABOOKS_IDENTIFIER,
        "Content-Type": "application/json"
    }
    payload = {
        "itemsku": skus,
        "opcode": "PUR"
    }

    for attempt in range(1, DANABOOKS_MAX_RETRIES + 1):
        resp = requests.post(DANABOOKS_URL, json=payload, headers=headers, timeout=30)

        if resp.status_code == 429:
            if attempt < DANABOOKS_MAX_RETRIES:
                print(f"[dana] 429 Rate limit, waiting {DANABOOKS_RETRY_WAIT}s (attempt {attempt}/{DANABOOKS_MAX_RETRIES})", flush=True)
                time.sleep(DANABOOKS_RETRY_WAIT)
                continue
            else:
                resp.raise_for_status()

        if resp.status_code == 520:
            if attempt < DANABOOKS_MAX_RETRIES:
                print(f"[dana] 520 Cloudflare error, waiting {DANABOOKS_RETRY_WAIT}s (attempt {attempt}/{DANABOOKS_MAX_RETRIES})", flush=True)
                time.sleep(DANABOOKS_RETRY_WAIT)
                continue
            else:
                print(f"[dana] 520 Cloudflare error after {DANABOOKS_MAX_RETRIES} attempts, skipping batch", flush=True)
                return {sku: None for sku in skus}

        resp.raise_for_status()
        data = resp.json()

        # Response: { "data": { "SKU1": [...records], "SKU2": [...records] } }
        sku_data = data.get("data", {})
        if not isinstance(sku_data, dict):
            return {sku: None for sku in skus}

        result = {}
        for sku in skus:
            records = sku_data.get(sku, [])
            if not records or not isinstance(records, list):
                result[sku] = None
                continue

            # Records are returned newest first — take the first one
            first_record = records[0]
            if not isinstance(first_record, dict):
                result[sku] = None
                continue

            price = first_record.get("item_price")
            if price is None:
                result[sku] = None
                continue

            try:
                result[sku] = float(price)
            except (ValueError, TypeError):
                print(f"[dana] WARNING: non-numeric item_price for {sku}: {price!r}", flush=True)
                result[sku] = None

        return result

    return {sku: None for sku in skus}


# ── Airtable helpers (Section 6)
# NOTE: renamed with danabooks_ prefix to avoid conflict with helpers of the
# same purpose already defined in Sections 2/3/4 ────────────────────────────

def danabooks_get_all_airtable_skus():
    """
    Fetch ALL records from French Inventories that have a SKU.
    Returns list of dicts: [{"record_id": ..., "sku": ..., "current_cost": ...}, ...]
    Retries each page up to 3 times on timeout/connection errors.
    """
    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{requests.utils.quote(DANABOOKS_TABLE_NAME)}"
    headers = {"Authorization": f"Bearer {AIRTABLE_API_KEY}"}
    params = {
        "filterByFormula": "{SKU}!=''",
        "fields[]": ["SKU", "Cost"],
        "pageSize": 100
    }

    records = []
    offset = None

    while True:
        if offset:
            params["offset"] = offset

        resp = None
        for attempt in range(1, 4):
            try:
                resp = requests.get(url, headers=headers, params=params, timeout=30)
                resp.raise_for_status()
                break
            except requests.exceptions.RequestException as e:
                print(f"[airtable-fetch] Attempt {attempt}/3 failed: {e}", flush=True)
                if attempt < 3:
                    time.sleep(5)
                else:
                    raise

        data = resp.json()

        for rec in data.get("records", []):
            sku = rec.get("fields", {}).get("SKU")
            if sku:
                current_cost = rec.get("fields", {}).get("Cost")
                if current_cost is not None:
                    try:
                        current_cost = float(current_cost)
                    except (ValueError, TypeError):
                        current_cost = None
                records.append({
                    "record_id": rec["id"],
                    "sku": sku,
                    "current_cost": current_cost
                })

        offset = data.get("offset")
        if not offset:
            break

    return records


def danabooks_update_airtable_cost(record_id, cost):
    """Update the Cost field in Airtable for a given record ID."""
    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{requests.utils.quote(DANABOOKS_TABLE_NAME)}/{record_id}"
    headers = {
        "Authorization": f"Bearer {AIRTABLE_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {"fields": {"Cost": cost}}
    resp = requests.patch(url, json=payload, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.json()


# ── Core sync job (Section 6) ──────────────────────────────────────────────────

def run_danabooks_auto_sync():
    # NOTE: renamed from run_auto_sync to run_danabooks_auto_sync to avoid
    # ambiguity alongside the other background sync jobs defined earlier.
    """
    Scheduled job:
    1. Fetch ALL SKUs from Airtable French Inventories
    2. For each batch of DANABOOKS_BATCH_SIZE SKUs, query Dana Books for latest purchase prices
    3. Only update Airtable if Cost is empty OR Dana Books price has changed
    """
    try:
        _run_danabooks_auto_sync_inner()
    except Exception as e:
        import traceback
        print(f"[auto-sync] FATAL UNCAUGHT ERROR: {e}", flush=True)
        print(traceback.format_exc(), flush=True)


def _run_danabooks_auto_sync_inner():
    print("[auto-sync] Starting scheduled cost sync...", flush=True)

    if not DANABOOKS_TOKEN:
        print("[auto-sync] ERROR: DANABOOKS_TOKEN is not set. Aborting.", flush=True)
        return

    try:
        all_skus = danabooks_get_all_airtable_skus()
    except Exception as e:
        print(f"[auto-sync] ERROR fetching Airtable SKUs: {e}", flush=True)
        return

    total = len(all_skus)
    print(f"[auto-sync] Total SKUs to check: {total}", flush=True)
    print(f"[auto-sync] Batch size: {DANABOOKS_BATCH_SIZE} SKUs per request", flush=True)

    updated = 0
    skipped_no_purchase = 0
    skipped_no_change = 0
    errors = 0
    done = 0

    # Build a lookup dict for quick access: sku -> item
    sku_map = {item["sku"]: item for item in all_skus}

    # Process in batches
    sku_list = list(sku_map.keys())
    for i in range(0, len(sku_list), DANABOOKS_BATCH_SIZE):
        batch_skus = sku_list[i:i + DANABOOKS_BATCH_SIZE]

        try:
            prices = get_prices_for_batch(batch_skus)
        except Exception as e:
            print(f"[auto-sync] ERROR fetching batch {i//DANABOOKS_BATCH_SIZE + 1}: {type(e).__name__}: {e}", flush=True)
            errors += len(batch_skus)
            done += len(batch_skus)
            continue

        for sku in batch_skus:
            item = sku_map[sku]
            record_id = item["record_id"]
            current_cost = item["current_cost"]
            dana_price = prices.get(sku)

            try:
                if dana_price is None:
                    skipped_no_purchase += 1
                elif current_cost is not None and current_cost == dana_price:
                    skipped_no_change += 1
                else:
                    danabooks_update_airtable_cost(record_id, dana_price)
                    print(
                        f"[auto-sync] UPDATED {sku} | "
                        f"old={current_cost if current_cost is not None else 'empty'} → new={dana_price}",
                        flush=True
                    )
                    updated += 1
            except Exception as e:
                print(f"[auto-sync] ERROR updating {sku}: {type(e).__name__}: {e}", flush=True)
                errors += 1

            done += 1

        if done % 500 == 0 or done == total:
            print(f"[auto-sync] Progress: {done}/{total} checked...", flush=True)

        # Delay between batch requests
        time.sleep(DANABOOKS_BATCH_DELAY)

    print(
        f"[auto-sync] Done. Updated={updated} | "
        f"No purchase in Dana Books={skipped_no_purchase} | "
        f"Price unchanged={skipped_no_change} | "
        f"Errors={errors}",
        flush=True
    )


# ── Scheduler — 9 AM, 2 PM, 8 PM IST (Section 6) ───────────────────────────────

danabooks_scheduler = BackgroundScheduler(timezone=DANABOOKS_IST)

danabooks_scheduler.add_job(
    run_danabooks_auto_sync,
    trigger=CronTrigger(hour=9, minute=0, timezone=DANABOOKS_IST),
    id="danabooks_sync_9am",
    name="Cost sync 9 AM IST"
)
danabooks_scheduler.add_job(
    run_danabooks_auto_sync,
    trigger=CronTrigger(hour=14, minute=0, timezone=DANABOOKS_IST),
    id="danabooks_sync_2pm",
    name="Cost sync 2 PM IST"
)
danabooks_scheduler.add_job(
    run_danabooks_auto_sync,
    trigger=CronTrigger(hour=20, minute=0, timezone=DANABOOKS_IST),
    id="danabooks_sync_8pm",
    name="Cost sync 8 PM IST"
)

danabooks_scheduler.start()
print("[scheduler] Dana Books cost sync scheduled at 9 AM, 2 PM, 8 PM IST", flush=True)


# ── Routes (Section 6) ─────────────────────────────────────────────────────────

@app.route("/sync-cost", methods=["POST"])
def sync_cost_manual():
    """
    Manually trigger cost sync for one or more specific SKUs.
    Body: { "skus": ["DNG1024", "DNG1025"] }
    Or single: { "sku": "DNG1024" }
    """
    body = request.get_json(force=True) or {}

    if "skus" in body:
        skus = body["skus"]
    elif "sku" in body:
        skus = [body["sku"]]
    else:
        return jsonify({"error": "Provide 'sku' or 'skus' in request body"}), 400

    results = []
    try:
        prices = get_prices_for_batch(skus)
    except Exception as e:
        return jsonify({"error": f"Dana Books API error: {type(e).__name__}: {e}"}), 500

    for sku in skus:
        result = {"sku": sku}
        try:
            dana_price = prices.get(sku)
            if dana_price is None:
                result["status"] = "skipped"
                result["reason"] = "No purchase records found in Dana Books"
                results.append(result)
                continue

            result["dana_price"] = dana_price

            url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{requests.utils.quote(DANABOOKS_TABLE_NAME)}"
            headers = {"Authorization": f"Bearer {AIRTABLE_API_KEY}"}
            params = {
                "filterByFormula": f"{{SKU}}='{sku}'",
                "maxRecords": 1,
                "fields[]": ["SKU", "Cost"]
            }
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            resp.raise_for_status()
            records = resp.json().get("records", [])

            if not records:
                result["status"] = "skipped"
                result["reason"] = "SKU not found in Airtable French Inventories"
                results.append(result)
                continue

            record_id = records[0]["id"]
            current_cost = records[0].get("fields", {}).get("Cost")

            if current_cost is not None:
                try:
                    current_cost = float(current_cost)
                except (ValueError, TypeError):
                    current_cost = None

            result["previous_cost"] = current_cost

            if current_cost is not None and current_cost == dana_price:
                result["status"] = "skipped"
                result["reason"] = "Price unchanged"
                results.append(result)
                continue

            danabooks_update_airtable_cost(record_id, dana_price)
            result["status"] = "updated"
            result["new_cost"] = dana_price

        except Exception as e:
            result["status"] = "error"
            result["reason"] = f"{type(e).__name__}: {e}"

        results.append(result)
        print(f"[manual-sync] {result}", flush=True)

    return jsonify({"results": results}), 200


@app.route("/danabooks/sync-all", methods=["POST"])
# NOTE: renamed from /sync-all to /danabooks/sync-all to avoid conflict with Section 2's /sync-all
def danabooks_sync_all_now():
    """Manually trigger the full Dana Books auto sync job immediately."""
    threading.Thread(target=run_danabooks_auto_sync, daemon=True).start()
    return jsonify({"message": "Full Dana Books sync started in background"}), 200


@app.route("/danabooks/health", methods=["GET"])
# NOTE: renamed from /health to /danabooks/health to avoid conflict with Section 4
def danabooks_health():
    jobs = [
        {"id": j.id, "name": j.name, "next_run": str(j.next_run_time)}
        for j in danabooks_scheduler.get_jobs()
    ]
    return jsonify({"status": "ok", "service": "danabooks-airtable-sync", "scheduled_jobs": jobs}), 200


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — SHOPIFY WEBHOOKS → AIRTABLE (orders / fulfillments / cancellations)
# ══════════════════════════════════════════════════════════════════════════════

SHOPIFY_TOKEN = os.getenv("SHOPIFY_TOKEN") or SHOPIFY_ACCESS_TOKEN
# 

print("🚀 Section 7 — Shopify Webhook ↔ Airtable Integration Starting...", flush=True)

# ---------------- BACKGROUND SYNC STATE (Section 7) ----------------
_shopint_sync_running = False
_shopint_sync_lock    = threading.Lock()

# ---------------- SHOPIFY → AIRTABLE PAYMENT STATUS MAP (Section 7) ----------------
SHOPINT_PAYMENT_STATUS_MAP = {
    "paid":               "Paid",
    "pending":            "Pending",
    "partially_paid":     "Pending",
    "refunded":           "Refunded",
    "voided":             "Cancelled",
    "partially_refunded": "Refunded",
    "authorized":         "Pending",
}

# ---------------- SHOPIFY → AIRTABLE SHIPPING STATUS LOGIC (Section 7) ----------------
SHOPINT_SHIPPED_STATUSES = {
    "label_printed",
    "label_purchased",
    "attempted_delivery",
    "ready_for_pickup",
    "confirmed",
    "in_transit",
    "out_for_delivery",
}

def shopint_determine_shipping_status_from_order(order):
    # Check if order is cancelled
    if order.get("cancelled_at"):
        return "Cancelled"

    fulfillments = order.get("fulfillments") or []
    fulfillment_status = (order.get("fulfillment_status") or "").lower()

    has_delivered = False
    has_shipped   = False
    has_fulfilled = False

    for f in fulfillments:
        shipment_status = (f.get("shipment_status") or "").lower()
        f_status        = (f.get("status") or "").lower()

        if shipment_status == "delivered":
            has_delivered = True
        elif shipment_status in SHOPINT_SHIPPED_STATUSES:
            has_shipped = True
        elif f_status == "success":
            has_fulfilled = True

    if has_delivered:
        return "Delivered"
    if has_shipped:
        return "Shipped"
    if has_fulfilled or fulfillment_status in ("fulfilled", "partial"):
        return "Fulfilled"
    return "New"


def shopint_determine_shipping_status_from_fulfillment(fulfillment):
    shipment_status = (fulfillment.get("shipment_status") or "").lower()
    f_status        = (fulfillment.get("status") or "").lower()

    if shipment_status == "delivered":
        return "Delivered"
    if shipment_status in SHOPINT_SHIPPED_STATUSES:
        return "Shipped"
    if f_status == "success":
        return "Fulfilled"
    return "Fulfilled"


# ---------------- SECURITY (Section 7) ----------------
def shopint_verify_webhook(data, hmac_header):
    # NOTE: renamed from verify_webhook to shopint_verify_webhook to avoid
    # conflict with Section 4's verify_webhook (different signature/purpose)
    if not hmac_header or not SHOPIFY_SECRET:
        return False
    digest = hmac.new(
        SHOPIFY_SECRET.encode("utf-8"),
        data,
        hashlib.sha256
    ).digest()
    computed_hmac = base64.b64encode(digest).decode("utf-8")
    return hmac.compare_digest(computed_hmac, hmac_header)


# ---------------- AIRTABLE HELPERS (Section 7) ----------------
def shopint_find_customer(phone, email):
    # NOTE: renamed from find_customer to shopint_find_customer to avoid
    # conflict with Section 4's find_customer (different signature/behavior)
    if phone:
        formula = f"{{Whatsapp number}}='{phone}'"
    elif email:
        formula = f"{{Mail id}}='{email}'"
    else:
        return None

    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{CUSTOMERS_TABLE_ID}"
    r = requests.get(url, headers=AIRTABLE_HEADERS, params={"filterByFormula": formula})
    data = r.json()

    if data.get("records"):
        return data["records"][0]["id"]
    return None


def shopint_create_customer(customer):
    # NOTE: renamed from create_customer to shopint_create_customer to avoid
    # conflict with Section 4's create_customer (different signature)
    payload = {
        "fields": {
            "Customer Name":          customer["name"],
            "Mail id":                customer.get("email"),
            "Whatsapp number":        customer.get("phone"),
            "Address":                customer.get("address"),
            "Acquired sales channel": "Shopify"
        }
    }
    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{CUSTOMERS_TABLE_ID}"
    r = requests.post(url, headers=AIRTABLE_HEADERS, json=payload)
    print(f"👤 Customer create status: {r.status_code}", flush=True)
    print(f"👤 Customer create response: {r.text}", flush=True)
    return r.json().get("id")


def shopint_find_sku_record(sku):
    if not sku:
        return None
    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{FRENCH_INVENTORIES_TABLE_ID}"
    r = requests.get(
        url,
        headers=AIRTABLE_HEADERS,
        params={"filterByFormula": f"{{SKU}}='{sku}'"}
    )
    data = r.json()
    if data.get("records"):
        return data["records"][0]["id"]
    return None


# ---------------- DUPLICATE CHECK (Section 7) ----------------
def shopint_order_exists(order_id):
    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDERS_TABLE_ID}"
    r = requests.get(
        url,
        headers=AIRTABLE_HEADERS,
        params={"filterByFormula": f"{{Order ID}}='{order_id}'"}
    )
    records = r.json().get("records", [])
    if records:
        return records[0]["id"]
    return None


# ---------------- UPDATE EXISTING ORDER STATUSES (Section 7) ----------------
def shopint_refresh_existing_order_statuses(order):
    order_id     = str(order["id"])
    order_number = order.get("name", "?")

    shipping_status = shopint_determine_shipping_status_from_order(order)
    shopify_payment = (order.get("financial_status") or "pending").lower()
    payment_status  = SHOPINT_PAYMENT_STATUS_MAP.get(shopify_payment, "Pending")

    # ── Extract Ship By ──
    ship_by = None
    fulfillments = order.get("fulfillments") or []
    if fulfillments:
        f = fulfillments[0]
        raw = (
            f.get("created_at") or
            f.get("updated_at") or
            f.get("shipped_at") or
            order.get("updated_at") or
            ""
        )
        if raw:
            ship_by = raw[:10]

    # --- Update Orders table ---
    orders_url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDERS_TABLE_ID}"
    r = requests.get(orders_url, headers=AIRTABLE_HEADERS,
                     params={"filterByFormula": f"{{Order ID}}='{order_id}'"})
    for record in r.json().get("records", []):
        fields = {"Shipping Status": shipping_status, "Payment Status": payment_status}
        if ship_by:
            fields["Ship By"] = ship_by
        requests.patch(f"{orders_url}/{record['id']}", headers=AIRTABLE_HEADERS, json={"fields": fields})

    # --- Update Order Line Items table ---
    line_url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDER_LINE_ITEMS_TABLE_ID}"
    r = requests.get(line_url, headers=AIRTABLE_HEADERS,
                     params={"filterByFormula": f"{{Order ID}}='{order_id}'"})
    line_records = r.json().get("records", [])
    for record in line_records:
        fields = {"Shipping Status": shipping_status, "Payment Status": payment_status}
        if ship_by:
            fields["Ship By"] = ship_by
        requests.patch(f"{line_url}/{record['id']}", headers=AIRTABLE_HEADERS, json={"fields": fields})

    print(f"🔄 {order_number} refreshed → Shipping: {shipping_status}, "
          f"Payment: {payment_status}, Ship By: {ship_by or 'N/A'} ({len(line_records)} line item(s))", flush=True)

# ---------------- ORDERS TABLE (Section 7) ----------------
def shopint_create_order_record(order, customer_id):
    order_date      = order["created_at"].split("T")[0]
    order_id        = str(order["id"])
    order_number    = order.get("name", "").replace("#", "")
    shopify_payment = (order.get("financial_status") or "pending").lower()
    payment_status  = SHOPINT_PAYMENT_STATUS_MAP.get(shopify_payment, "Pending")

    # add ship by
    ship_by = None
    if order.get("fulfillments"):
        fulfilled_at = order["fulfillments"][0].get("created_at", "")
        if fulfilled_at:
            ship_by = fulfilled_at[:10]

    fields = {
        "Order ID":        order_id,
        "Customer":        [customer_id],
        "Order Date":      order_date,
        "Sales Channel":   "Shopify",
        "Shipping Status": shopint_determine_shipping_status_from_order(order),
        "Payment Status":  payment_status,
    }
    if ship_by:
        fields["Ship By"] = ship_by

    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDERS_TABLE_ID}"
    r = requests.post(url, headers=AIRTABLE_HEADERS, json={"fields": fields})

    if r.status_code in (200, 201):
        order_record_id = r.json().get("id")
        print(f"✅ Created Orders record: {order_number} (Airtable ID: {order_record_id})", flush=True)
        return order_record_id
    else:
        print(f"❌ Failed to create Orders record: {r.status_code} — {r.text}", flush=True)
        return None


# ---------------- SHIPPING STATUS UPDATE (Section 7) ----------------
def shopint_update_shipping_status(order_id, status, ship_by=None):
    # --- Update Orders table ---
    orders_url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDERS_TABLE_ID}"
    r = requests.get(
        orders_url,
        headers=AIRTABLE_HEADERS,
        params={"filterByFormula": f"{{Order ID}}='{order_id}'"}
    )
    order_records = r.json().get("records", [])
    for record in order_records:
        fields = {"Shipping Status": status}
        if ship_by:
            fields["Ship By"] = ship_by
        requests.patch(
            f"{orders_url}/{record['id']}",
            headers=AIRTABLE_HEADERS,
            json={"fields": fields}
        )
    print(f"🚚 Orders table Shipping Status → '{status}'", flush=True)

    # --- Update Order Line Items table ---
    line_url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDER_LINE_ITEMS_TABLE_ID}"
    r = requests.get(
        line_url,
        headers=AIRTABLE_HEADERS,
        params={"filterByFormula": f"{{Order ID}}='{order_id}'"}
    )
    line_records = r.json().get("records", [])
    if not line_records:
        print(f"⚠️ No line items found for Order ID {order_id}", flush=True)
    for record in line_records:
        fields = {"Shipping Status": status}
        if ship_by:
            fields["Ship By"] = ship_by
        requests.patch(
            f"{line_url}/{record['id']}",
            headers=AIRTABLE_HEADERS,
            json={"fields": fields}
        )
    print(f"🚚 Order Line Items Shipping Status → '{status}' on {len(line_records)} row(s)", flush=True)

# ---------------- ORDER LINE ITEM CREATION (Section 7) ----------------
def shopint_create_order_line_items(order, customer_id, order_record_id):
    print("🧾 Creating order line item records...", flush=True)

    order_date     = order["created_at"].split("T")[0]
    order_id       = str(order["id"])
    order_number   = order.get("name", "").replace("#", "")
    shopify_status = order.get("financial_status", "pending").lower()
    payment_status = SHOPINT_PAYMENT_STATUS_MAP.get(shopify_status, "Pending")
    shipping_status = shopint_determine_shipping_status_from_order(order)

    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDER_LINE_ITEMS_TABLE_ID}"

    for line in order.get("line_items", []):
        sku        = line.get("sku")
        product_id = shopint_find_sku_record(sku)

        price     = float(line.get("price", 0))
        qty       = int(line.get("quantity", 1))

        tax_lines = line.get("tax_lines", [])
        tax_rate  = tax_lines[0].get("rate", 0) if tax_lines else 0
        tax_pct   = f"{int(tax_rate * 100)}%"

        fields = {
            "Order ID":        order_id,
            "Order Number":    order_number,
            "Customer":        [customer_id],
            "Order Date":      order_date,
            "Rate":            price,
            "Qty":             qty,
            "Tax Type":        "5%",
            "Payment Status":  payment_status,
            "Shipping Status": shipping_status,
            "Sales Channel":   "Shopify",
        }

        if order_record_id:
            fields["Order"] = [order_record_id]

        if product_id:
            fields["Product"] = [product_id]
        else:
            print(f"⚠️ SKU '{sku}' not found in French Inventories — Product field left empty", flush=True)

        r = requests.post(url, headers=AIRTABLE_HEADERS, json={"fields": fields})

        if r.status_code in (200, 201):
            print(f"✅ Created line item: '{line.get('title')}' (SKU: {sku})", flush=True)
        else:
            print(f"❌ Failed line item '{line.get('title')}': {r.status_code} — {r.text}", flush=True)


# ---------------- MAIN LOGIC (Section 7) ----------------
def shopint_process_order(order):
    # NOTE: renamed from process_order to shopint_process_order to avoid
    # conflict with Section 2's process_order (Amazon orders, different signature)
    order_id = str(order["id"])

    existing_order_record_id = shopint_order_exists(order_id)
    if existing_order_record_id:
        print(f"⏭️ Order {order_id} already exists in Orders table — skipping", flush=True)
        return

    customer    = order.get("customer") or {}
    customer_id = shopint_find_customer(customer.get("phone"), customer.get("email"))

    if not customer_id:
        customer_id = shopint_create_customer({
            "name":    f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip(),
            "email":   customer.get("email"),
            "phone":   customer.get("phone"),
            "address": (order.get("shipping_address") or {}).get("address1")
        })

    if not customer_id:
        print("❌ Could not find or create customer — aborting order", flush=True)
        return

    order_record_id = shopint_create_order_record(order, customer_id)
    if not order_record_id:
        print("❌ Could not create Orders record — aborting line items", flush=True)
        return

    shopint_create_order_line_items(order, customer_id, order_record_id)


# ---------------- WEBHOOK : ORDERS (Section 7) ----------------
@app.route("/shopify/webhook/orders", methods=["POST"])
def shopify_orders():
    data        = request.get_data()
    hmac_header = request.headers.get("X-Shopify-Hmac-Sha256")

    if not shopint_verify_webhook(data, hmac_header):
        return "Unauthorized", 401

    shopint_process_order(request.json)
    return jsonify({"status": "ok"})


# ---------------- WEBHOOK : FULFILLMENTS (Section 7) ----------------
@app.route("/shopify/webhook/fulfillments", methods=["POST"])
def shopify_fulfillments():
    data        = request.get_data()
    hmac_header = request.headers.get("X-Shopify-Hmac-Sha256")

    if not shopint_verify_webhook(data, hmac_header):
        return "Unauthorized", 401

    payload  = request.json
    order_id = payload.get("order_id")

    if not order_id:
        return jsonify({"status": "no order id"}), 200

    new_status   = shopint_determine_shipping_status_from_fulfillment(payload)
    fulfilled_at = (payload.get("created_at") or "")[:10]
    shopint_update_shipping_status(str(order_id), new_status, fulfilled_at)
    return jsonify({"status": new_status.lower()})


# ---------------- WEBHOOK : CANCELLATIONS (Section 7) ----------------
@app.route("/shopify/webhook/cancellations", methods=["POST"])
def shopify_cancellations():
    data        = request.get_data()
    hmac_header = request.headers.get("X-Shopify-Hmac-Sha256")

    if not shopint_verify_webhook(data, hmac_header):
        return "Unauthorized", 401

    payload    = request.json
    order_id   = str(payload.get("id", ""))
    order_name = payload.get("name", "?")

    if not order_id:
        return jsonify({"status": "no order id"}), 200

    shopint_update_shipping_status(order_id, "Cancelled")
    print(f"🚫 Order {order_name} marked as Cancelled", flush=True)
    return jsonify({"status": "cancelled"})


# ---------------- SYNC ALL SHOPIFY ORDERS (Section 7) ----------------
def shopint_fetch_all_shopify_orders():
    # NOTE: fixed to use SHOPIFY_STORE (already the full *.myshopify.com
    # domain) and the shared API_VERSION directly — see the fix note at the
    # top of this section for why the original ".myshopify.com" suffix was removed.
    orders = []
    url    = f"https://{SHOPIFY_STORE}/admin/api/{API_VERSION}/orders.json"
    params = {"limit": 250, "status": "any"}

    while url:
        r     = requests.get(url, headers={"X-Shopify-Access-Token": SHOPIFY_TOKEN}, params=params)
        batch = r.json().get("orders", [])
        orders.extend(batch)
        print(f"📦 Fetched {len(batch)} orders (total: {len(orders)})", flush=True)

        link   = r.headers.get("Link", "")
        url    = None
        params = {}
        if 'rel="next"' in link:
            for part in link.split(","):
                if 'rel="next"' in part:
                    url = part.split(";")[0].strip().strip("<>")
                    break

    return orders


def _shopint_do_full_sync():
    global _shopint_sync_running

    try:
        all_orders = shopint_fetch_all_shopify_orders()
        print(f"✅ Total orders from Shopify: {len(all_orders)}", flush=True)

        synced  = 0
        updated = 0
        failed  = 0

        for order in all_orders:
            order_name = order.get("name", "?")
            try:
                order_id = str(order["id"])
                if shopint_order_exists(order_id):
                    shopint_refresh_existing_order_statuses(order)
                    updated += 1
                    continue

                customer    = order.get("customer") or {}
                customer_id = shopint_find_customer(customer.get("phone"), customer.get("email"))

                if not customer_id:
                    customer_id = shopint_create_customer({
                        "name":    f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip(),
                        "email":   customer.get("email"),
                        "phone":   customer.get("phone"),
                        "address": (order.get("shipping_address") or {}).get("address1")
                    })

                if not customer_id:
                    print(f"❌ {order_name} — could not find/create customer", flush=True)
                    failed += 1
                    continue

                order_record_id = shopint_create_order_record(order, customer_id)
                if not order_record_id:
                    failed += 1
                    continue

                shopint_create_order_line_items(order, customer_id, order_record_id)
                print(f"✅ {order_name} synced", flush=True)
                synced += 1

            except Exception as e:
                print(f"❌ {order_name} — error: {e}", flush=True)
                failed += 1

        print(
            f"🎉 Sync complete: total={len(all_orders)} synced={synced} "
            f"updated={updated} failed={failed}",
            flush=True
        )

    except Exception as e:
        print(f"❌ Background sync crashed: {e}", flush=True)

    finally:
        with _shopint_sync_lock:
            _shopint_sync_running = False


# ---------------- WEBHOOK : ORDER UPDATED (Section 7) ----------------
@app.route("/shopify/webhook/order-updated", methods=["POST"])
def shopify_order_updated():
    data        = request.get_data()
    hmac_header = request.headers.get("X-Shopify-Hmac-Sha256")

    if not shopint_verify_webhook(data, hmac_header):
        return "Unauthorized", 401

    order    = request.json
    order_id = str(order.get("id", ""))

    if not order_id:
        return jsonify({"status": "no order id"}), 200

    # Update payment status
    shopify_payment = (order.get("financial_status") or "pending").lower()
    payment_status  = SHOPINT_PAYMENT_STATUS_MAP.get(shopify_payment, "Pending")

    # Update shipping status
    shipping_status = shopint_determine_shipping_status_from_order(order)

    # ── Extract Ship By ──
    ship_by = None
    fulfillments = order.get("fulfillments") or []
    if fulfillments:
        f = fulfillments[0]
        raw = (
            f.get("created_at") or
            f.get("updated_at") or
            order.get("updated_at") or
            ""
        )
        if raw:
            ship_by = raw[:10]

    # --- Update Orders table ---
    orders_url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDERS_TABLE_ID}"
    r = requests.get(
        orders_url,
        headers=AIRTABLE_HEADERS,
        params={"filterByFormula": f"{{Order ID}}='{order_id}'"}
    )
    for record in r.json().get("records", []):
        requests.patch(
            f"{orders_url}/{record['id']}",
            headers=AIRTABLE_HEADERS,
            json={"fields": {
                "Shipping Status": shipping_status,
                "Payment Status":  payment_status,
                **( {"Ship By": ship_by} if ship_by else {} )
            }}
        )

    # --- Update Order Line Items table ---
    line_url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{ORDER_LINE_ITEMS_TABLE_ID}"
    r = requests.get(
        line_url,
        headers=AIRTABLE_HEADERS,
        params={"filterByFormula": f"{{Order ID}}='{order_id}'"}
    )
    for record in r.json().get("records", []):
        requests.patch(
            f"{line_url}/{record['id']}",
            headers=AIRTABLE_HEADERS,
            json={"fields": {
                "Shipping Status": shipping_status,
                "Payment Status":  payment_status,
                **( {"Ship By": ship_by} if ship_by else {} )
            }}
        )

    print(f"🔄 Order {order.get('name','?')} updated → Payment: {payment_status}, Shipping: {shipping_status}", flush=True)
    return jsonify({"status": "ok", "payment": payment_status, "shipping": shipping_status})


@app.route("/shopify-integration/sync", methods=["GET"])
# NOTE: namespaced from /sync to /shopify-integration/sync to keep it distinct
# from the other sync routes already defined (/sync-all, /sync-cost, etc.)
def shopint_sync_all_orders():
    global _shopint_sync_running

    if not SHOPIFY_STORE or not SHOPIFY_TOKEN:
        return jsonify({
            "status":  "error",
            "message": "SHOPIFY_STORE or SHOPIFY_TOKEN env variable not set in Render"
        }), 500

    with _shopint_sync_lock:
        if _shopint_sync_running:
            return jsonify({
                "status":  "already_running",
                "message": "A sync is already running. Watch Render logs for progress."
            }), 409
        _shopint_sync_running = True

    print("🔄 Manual sync triggered — running in background...", flush=True)
    threading.Thread(target=_shopint_do_full_sync, daemon=True).start()

    return jsonify({
        "status":  "started",
        "message": "Sync started in background. Watch Render logs for progress. Look for '🎉 Sync complete' when done."
    }), 202


# ---------------- HEALTH CHECK (Section 7) ----------------
@app.route("/shopify-integration/health", methods=["GET"])
# NOTE: renamed from /health to /shopify-integration/health to avoid conflict with Section 4
def shopify_integration_health():
    return jsonify({"status": "ok", "service": "shopify-airtable-integration"}), 200


# ══════════════════════════════════════════════════════════════════════════════
# RUN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)