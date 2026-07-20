"""
session_report.py — Section 5: Daily Session Report → Slack (#session)
Split out of the original combined app.py. Code below is unchanged from the
merged version; only the file location and these import lines changed.

NOTE: the requested filename was "session-report.py" but Python module names
cannot contain hyphens (import session-report is invalid syntax), so this is
named session_report.py instead. Route paths and behavior are unchanged.
"""
import os
import io
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from flask import jsonify
import requests
import openpyxl
from apscheduler.schedulers.background import BackgroundScheduler

from shared import app
from delivery_tracker import SHOPIFY_STORE, SHOPIFY_ACCESS_TOKEN

# NOTE: reuses SHOPIFY_STORE / SHOPIFY_ACCESS_TOKEN already defined in Section 1.
# Falls back to SHOPIFY_ADMIN_API_TOKEN if that's the env var name configured
# for this particular service on Render.

SESSION_REPORT_SHOPIFY_TOKEN = os.getenv("SHOPIFY_ADMIN_API_TOKEN") or SHOPIFY_ACCESS_TOKEN
SLACK_BOT_TOKEN               = os.getenv("SLACK_BOT_TOKEN")        # xoxb-...
SESSION_REPORT_CHANNEL_ID     = os.getenv("SLACK_CHANNEL_ID", "C0B9V9U312L")

IST = ZoneInfo("Asia/Kolkata")


# ── SHOPIFY: fetch sessions by landing page (Section 5) ───────────────────────

def fetch_sessions(date_str):
    print(f"[fetch_sessions] Querying Shopify for date: {date_str}", flush=True)

    url = f"https://{SHOPIFY_STORE}/admin/api/2026-04/graphql.json"
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": SESSION_REPORT_SHOPIFY_TOKEN,
    }

    query = f"""
    {{
      shopifyqlQuery(query: "FROM sessions SHOW landing_page_type, landing_page_path, online_store_visitors, sessions SINCE {date_str} UNTIL {date_str} GROUP BY landing_page_type, landing_page_path ORDER BY sessions DESC") {{
        tableData {{
          columns {{ name }}
          rows
        }}
        parseErrors
      }}
    }}
    """

    print(f"[fetch_sessions] Sending ShopifyQL request...", flush=True)
    resp = requests.post(url, json={"query": query}, headers=headers, timeout=30)
    print(f"[fetch_sessions] Response status: {resp.status_code}", flush=True)
    data = resp.json()

    if "errors" in data:
        print(f"[fetch_sessions] GraphQL errors: {data['errors']}", flush=True)
        raise RuntimeError(f"GraphQL errors: {data['errors']}")

    shopify_data = data.get("data", {}).get("shopifyqlQuery", {})
    parse_errors = shopify_data.get("parseErrors")
    if parse_errors:
        print(f"[fetch_sessions] ShopifyQL parse errors: {parse_errors}", flush=True)
        raise ValueError(f"ShopifyQL parse errors: {parse_errors}")

    table = shopify_data.get("tableData", {})
    if not table:
        print(f"[fetch_sessions] WARNING - No tableData returned.", flush=True)
        return []

    columns = [col["name"] for col in table.get("columns", [])]
    rows    = table.get("rows", [])
    print(f"[fetch_sessions] Columns: {columns}", flush=True)
    print(f"[fetch_sessions] Total rows: {len(rows)}", flush=True)

    results = []
    for row in rows:
        if isinstance(row, list):
            record = dict(zip(columns, row))
        elif isinstance(row, dict):
            record = row
        else:
            continue
        results.append(record)

    print(f"[fetch_sessions] Successfully parsed {len(results)} rows.", flush=True)
    return results


# ── BUILD EXCEL (Section 5) ────────────────────────────────────────────────────

def build_excel(rows, date_str):
    """
    Filters Product rows, keeps 3 columns, returns Excel bytes.
    """
    print(f"[build_excel] Starting Excel build for {date_str}. Input rows: {len(rows)}", flush=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessions by Landing Page"

    # Header
    headers = ["Landing page path", "Online store visitors", "Sessions"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows — only Product type
    count = 0
    skipped = 0
    for r in rows:
        page_type = str(r.get("landing_page_type", "")).strip().lower()
        if page_type != "product":
            print(f"[build_excel] Skipping row — type='{page_type}', path='{r.get('landing_page_path', '')}'", flush=True)
            skipped += 1
            continue
        ws.append([
            r.get("landing_page_path", ""),
            r.get("online_store_visitors", 0),
            r.get("sessions", 0),
        ])
        count += 1

    print(f"[build_excel] Rows added (Product): {count} | Rows skipped (non-Product): {skipped}", flush=True)

    # Column widths
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_size_kb = round(len(buf.getvalue()) / 1024, 2)
    print(f"[build_excel] Excel file built successfully. Size: {excel_size_kb} KB", flush=True)
    return buf.getvalue(), count


# ── SEND TO SLACK (Section 5) ─────────────────────────────────────────────────

def send_session_report_to_slack(excel_bytes, date_str, row_count):
    # NOTE: renamed from send_to_slack to send_session_report_to_slack
    # to keep this service's Slack upload flow distinct from other sections.
    filename = f"sessions_by_landing_page_{date_str}.xlsx"
    file_size = len(excel_bytes)
    print(f"[send_to_slack] Preparing to upload '{filename}' to Slack channel {SESSION_REPORT_CHANNEL_ID}", flush=True)
    print(f"[send_to_slack] File size: {round(file_size/1024, 2)} KB | Product rows: {row_count}", flush=True)

    # Step 1 — Get upload URL
    print(f"[send_to_slack] Step 1: Requesting upload URL from Slack...", flush=True)
    url_resp = requests.post(
        "https://slack.com/api/files.getUploadURLExternal",
        headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"},
        data={"filename": filename, "length": file_size},
        timeout=30,
    )
    url_resp.raise_for_status()
    url_result = url_resp.json()
    print(f"[send_to_slack] Step 1 response: ok={url_result.get('ok')} error={url_result.get('error','none')}", flush=True)

    if not url_result.get("ok"):
        raise RuntimeError(f"Slack getUploadURLExternal failed: {url_result.get('error')}")

    upload_url = url_result["upload_url"]
    file_id    = url_result["file_id"]
    print(f"[send_to_slack] Got upload URL. File ID: {file_id}", flush=True)

    # Step 2 — Upload file bytes to the upload URL
    print(f"[send_to_slack] Step 2: Uploading file bytes...", flush=True)
    upload_resp = requests.post(
        upload_url,
        files={"file": (filename, excel_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=60,
    )
    upload_resp.raise_for_status()
    print(f"[send_to_slack] Step 2 done. HTTP status: {upload_resp.status_code}", flush=True)

    # Step 3 — Complete upload and share to channel
    print(f"[send_to_slack] Step 3: Completing upload and posting to channel...", flush=True)
    complete_resp = requests.post(
        "https://slack.com/api/files.completeUploadExternal",
        headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"},
        json={
            "files": [{"id": file_id, "title": filename}],
            "channel_id": SESSION_REPORT_CHANNEL_ID,
            "initial_comment": (
                f":bar_chart: *Sessions by Landing Page — {date_str}*\n"
                f"Product pages: *{row_count} rows* | Columns: Landing page path, Online store visitors, Sessions"
            ),
        },
        timeout=30,
    )
    complete_resp.raise_for_status()
    complete_result = complete_resp.json()
    print(f"[send_to_slack] Step 3 response: ok={complete_result.get('ok')} error={complete_result.get('error','none')}", flush=True)

    if not complete_result.get("ok"):
        raise RuntimeError(f"Slack completeUploadExternal failed: {complete_result.get('error')}")

    print(f"[send_to_slack] SUCCESS - File posted to Slack channel. File ID: {file_id}", flush=True)
    print(f"[session-slack] Sent {row_count} product rows for {date_str} to Slack.", flush=True)


# ── MAIN JOB (Section 5) ───────────────────────────────────────────────────────

def run_session_report_job():
    # NOTE: renamed from run_job to run_session_report_job to avoid ambiguity
    # alongside the other background jobs defined in earlier sections.
    print(f"[run_job] ─────────────────────────────────────────", flush=True)
    print(f"[run_job] Job started at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    try:
        yesterday = (datetime.now(IST) - timedelta(days=1)).strftime("%Y-%m-%d")
        print(f"[run_job] Target date (yesterday): {yesterday}", flush=True)

        rows = fetch_sessions(yesterday)
        print(f"[run_job] Got {len(rows)} total rows from Shopify.", flush=True)

        excel_bytes, count = build_excel(rows, yesterday)
        print(f"[run_job] Filtered to {count} Product rows.", flush=True)

        send_session_report_to_slack(excel_bytes, yesterday, count)
        print(f"[run_job] Job completed successfully.", flush=True)

    except Exception as e:
        print(f"[run_job] ERROR: {e}", flush=True)
    print(f"[run_job] ─────────────────────────────────────────", flush=True)


# ── Routes (Section 5) ─────────────────────────────────────────────────────────

@app.route("/session-report/health", methods=["GET"])
# NOTE: renamed from /health to /session-report/health to avoid conflict with Section 4
def session_report_health():
    print(f"[session-report/health] Health check called at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    return jsonify({"status": "ok", "service": "session-report-slack"})


@app.route("/session-report/run-now", methods=["GET"])
# NOTE: renamed from /run-now to /session-report/run-now for a unique, namespaced path
def session_report_run_now():
    """Trigger the job manually for testing."""
    print(f"[session-report/run-now] Manual trigger called at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    run_session_report_job()
    return jsonify({"status": "done"})


@app.route("/session-report/run-date/<date_str>", methods=["GET"])
# NOTE: renamed from /run-date/<date_str> to /session-report/run-date/<date_str>
def session_report_run_date(date_str):
    """Fetch a specific date. Format: YYYY-MM-DD  e.g. /session-report/run-date/2026-06-07"""
    print(f"[session-report/run-date] Manual date trigger called for: {date_str}", flush=True)
    try:
        rows = fetch_sessions(date_str)
        excel_bytes, count = build_excel(rows, date_str)
        send_session_report_to_slack(excel_bytes, date_str, count)
        print(f"[session-report/run-date] Completed for {date_str}. Product rows: {count}", flush=True)
        return jsonify({"status": "done", "product_rows": count, "date": date_str})
    except Exception as e:
        print(f"[session-report/run-date] ERROR for {date_str}: {e}", flush=True)
        return jsonify({"status": "error", "message": str(e)}), 500


# ── Scheduler (Section 5) ──────────────────────────────────────────────────────
session_report_scheduler = BackgroundScheduler(timezone=IST)
session_report_scheduler.add_job(run_session_report_job, "cron", hour=11, minute=0)
session_report_scheduler.start()
print(f"[startup] Session report scheduler started. Job will run daily at 17:00 IST.", flush=True)
print(f"[startup] SESSION REPORT SHOPIFY_STORE     : {SHOPIFY_STORE}", flush=True)
print(f"[startup] SESSION REPORT SHOPIFY_TOKEN set : {'YES' if SESSION_REPORT_SHOPIFY_TOKEN else 'NO ⚠️'}", flush=True)
print(f"[startup] SESSION REPORT SLACK_TOKEN set   : {'YES' if SLACK_BOT_TOKEN else 'NO ⚠️'}", flush=True)
print(f"[startup] SESSION REPORT SLACK_CHANNEL_ID  : {SESSION_REPORT_CHANNEL_ID}", flush=True)